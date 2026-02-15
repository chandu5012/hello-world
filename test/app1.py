from flask import Flask, render_template, request, jsonify, Response, send_file
from flask_socketio import SocketIO, emit
import json
import time
import paramiko
import shlex
import threading
import uuid
import os
from io import StringIO, BytesIO
from queue import Queue
from collections import defaultdict
from datetime import datetime
import tempfile

# Excel parsing
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max upload
socketio = SocketIO(app, cors_allowed_origins="*")

# ============================================================
# RUN REGISTRY - In-memory storage for execution state
# ============================================================
run_registry = {}
# Structure:
# {
#   run_id: {
#     'status': 'PENDING' | 'RUNNING' | 'SUCCESS' | 'FAILED' | 'STOPPED' | 'PARTIAL_SUCCESS',
#     'logs': [],
#     'tab': str,
#     'params': dict,
#     'thread': Thread,
#     'ssh_client': SSHClient or None,
#     'stop_requested': bool,
#     'rows_status': [],  # For multi-table runs: [{row_index, table, keys, status, error}]
#     'total_rows': 0,
#     'started_at': datetime,
#     'finished_at': datetime
#   }
# }

# Tab to script mapping
TAB_SCRIPT_MAP = {
    'data_comparison': 'compare_job.sh',  # Script name only (not path) - for single_table mode
    'data_load': 'run_data_load.sh',
    'schema_generation': 'run_schema_generation.sh',
    'file_load': 'run_file_load.sh',
    'file_download': 'run_file_download.sh'
}

# Multi-table script (used only when mode == 'multiple_tables')
MULTI_TABLE_SCRIPT = 'run_data_comparison.sh'

# Single-table script for data comparison
SINGLE_TABLE_SCRIPT = 'compare_job.sh'

# ============================================================
# GLOBAL WORKING DIRECTORY MANAGEMENT
# ============================================================
def resolve_execution_context(params, run_id=None):
    """
    Resolve execution context including working_dir, batch_id, and env.
    
    This is the CENTRAL function for computing working_dir.
    All execution paths MUST use this function to get execution context.
    
    Returns:
        dict: {
            'working_dir': str,  # e.g., '/home/chandra/BATCH001/dev'
            'batch_id': str,
            'env': str,
            'unix_config': dict,
            'error': str or None
        }
    """
    # Get Unix config from params
    unix_config = get_unix_config_from_params(params)
    
    # Check for missing host error
    if 'error' in unix_config:
        return {
            'working_dir': None,
            'batch_id': None,
            'env': None,
            'unix_config': unix_config,
            'error': unix_config['error']
        }
    
    # Validate batch_id is provided
    batch_id = unix_config.get('batch_id', '').strip()
    if not batch_id:
        return {
            'working_dir': None,
            'batch_id': None,
            'env': None,
            'unix_config': unix_config,
            'error': 'Batch ID is required for execution'
        }
    
    # Determine environment from unix host
    env, env_error = determine_environment_from_host(unix_config['host'])
    if env_error:
        return {
            'working_dir': None,
            'batch_id': batch_id,
            'env': None,
            'unix_config': unix_config,
            'error': env_error
        }
    
    # Compute working directory ONCE
    working_dir = f'/home/chandra/{batch_id}/{env}'
    
    return {
        'working_dir': working_dir,
        'batch_id': batch_id,
        'env': env,
        'unix_config': unix_config,
        'error': None
    }


def get_working_dir(run_id):
    """
    Get the global working_dir for a run from the registry.
    
    This ensures all execution paths use the same working_dir.
    """
    if run_id in run_registry:
        return run_registry[run_id].get('working_dir')
    return None


def set_execution_context(run_id, working_dir, batch_id, env):
    """
    Store execution context in run registry.
    Called once when execution starts.
    """
    if run_id in run_registry:
        run_registry[run_id]['working_dir'] = working_dir
        run_registry[run_id]['batch_id'] = batch_id
        run_registry[run_id]['env'] = env


def determine_environment_from_host(unix_host):
    """
    Determine environment (dev/sit/uat) from unix_host.
    
    Mapping:
    - compute-1 or compute-2 -> dev
    - compute-3 or compute-4 -> sit
    - compute-5 or compute-6 -> uat
    
    Returns: (env, error) tuple. If env cannot be determined, returns (None, error_message).
    """
    host_lower = unix_host.lower() if unix_host else ''
    
    if 'compute-1' in host_lower or 'compute-2' in host_lower:
        return ('dev', None)
    elif 'compute-3' in host_lower or 'compute-4' in host_lower:
        return ('sit', None)
    elif 'compute-5' in host_lower or 'compute-6' in host_lower:
        return ('uat', None)
    else:
        return (None, f"Unable to determine environment (dev/sit/uat) from unix host: {unix_host}")


def build_command_with_working_dir(script_path, params, batch_id, env):
    """
    Build shell command with cd to working directory and properly escaped arguments.
    
    Command format: cd /home/chandra/{batch_id}/{env}/ && {script_path} {args}
    """
    working_dir = f'/home/chandra/{batch_id}/{env}'
    
    # Build argument list
    command_parts = [script_path]
    
    for key, value in params.items():
        if value is not None and value != '':
            escaped_value = escape_shell_arg(value)
            # Convert camelCase to snake_case for CLI args
            cli_key = ''.join(['_' + c.lower() if c.isupper() else c for c in key]).lstrip('_')
            command_parts.append(f'--{cli_key}={escaped_value}')
    
    script_command = ' '.join(command_parts)
    
    # Build full command with cd
    full_command = f'cd {shlex.quote(working_dir)} && {script_command}'
    
    return full_command, working_dir


def normalize_comparison_mode(params):
    """
    Extract and normalize the comparison mode from request params.
    
    Accepts common names: mode, compare_mode, comparison_mode, table_mode
    Normalizes values like "Single Table", "single", "single_table" into "single_table"
    Normalizes "Multiple Tables", "multi", "multiple_tables" into "multiple_tables"
    
    Returns: 'single_table' (default) or 'multiple_tables'
    """
    # Extract mode from various possible field names
    raw_mode = (
        params.get('mode') or
        params.get('compare_mode') or
        params.get('comparison_mode') or
        params.get('table_mode') or
        params.get('tableMode') or
        ''
    ).strip().lower().replace(' ', '_').replace('-', '_')
    
    # Normalize to canonical values
    if raw_mode in ['multiple_tables', 'multi', 'multiple', 'multi_table']:
        return 'multiple_tables'
    
    # Default to single_table
    return 'single_table'


def normalize_comparison_type(params):
    """
    Extract and normalize the comparison type from request params.
    
    Returns: 'database_to_database', 'file_to_database', 'file_to_file', 'database_to_file'
    """
    raw_type = (
        params.get('comparison_type') or
        params.get('comparisonType') or
        params.get('comparison-type-select') or
        ''
    ).strip().lower().replace('-', '_')
    
    if raw_type in ['db_to_db', 'database_to_database', 'db_db']:
        return 'database_to_database'
    elif raw_type in ['file_to_db', 'file_to_database', 'file_db']:
        return 'file_to_database'
    elif raw_type in ['file_to_file', 'file_file']:
        return 'file_to_file'
    elif raw_type in ['db_to_file', 'database_to_file', 'db_file']:
        return 'database_to_file'
    
    # Default
    return 'database_to_database'


def determine_download_filename(params):
    """
    Determine the output filename for single_table comparison based on comparison type.
    
    Output filename rules:
    - database_to_database → {target_table}.csv
    - file_to_database → {target_table}.csv
    - file_to_file → file_to_file.csv
    - database_to_file → db_to_file.csv
    
    Returns dict with 'filename' and 'unix_path'
    """
    comparison_type = normalize_comparison_type(params)
    
    # Get target table name for db_to_db and file_to_db
    target_table = (
        params.get('target_table_name') or
        params.get('targetTableName') or
        params.get('comparisonTargetTableName') or
        params.get('target_table') or
        params.get('targetTable') or
        ''
    ).strip()
    
    if comparison_type == 'database_to_database':
        if target_table:
            filename = f'{target_table}.csv'
        else:
            filename = 'comparison_output.csv'
    elif comparison_type == 'file_to_database':
        if target_table:
            filename = f'{target_table}.csv'
        else:
            filename = 'comparison_output.csv'
    elif comparison_type == 'file_to_file':
        filename = 'file_to_file.csv'
    elif comparison_type == 'database_to_file':
        filename = 'db_to_file.csv'
    else:
        filename = 'comparison_output.csv'
    
    # Assume the script outputs to /tmp/output.csv or a known location
    # The download endpoint will try multiple paths if the primary fails
    unix_path = '/tmp/output.csv'
    
    return {
        'filename': filename,
        'unix_path': unix_path
    }


def extract_time_zone(params):
    """
    Extract time zone from request params.
    
    Returns the time zone value or 'N/A' as default.
    Accepts: time_zone, timeZone, comparisonTimeZone, loadTimeZone
    """
    time_zone = (
        params.get('time_zone') or
        params.get('timeZone') or
        params.get('comparisonTimeZone') or
        params.get('comparison_time_zone') or
        params.get('loadTimeZone') or
        params.get('load_time_zone') or
        ''
    ).strip()
    
    return time_zone if time_zone else 'N/A'


def extract_data_load_fields(params):
    """
    Extract Data Load specific fields from request params.
    
    Returns dict with:
    - target_load_type: 'overwrite' or 'append'
    - hadoop_path: Hadoop path (required if target is Hive)
    - partition_id: Partition ID (defaults to 'N/A' if empty)
    - target_db_type: Target database type for Hive detection
    """
    result = {}
    
    # Target Load Type (overwrite/append)
    target_load_type = (
        params.get('target_load_type') or
        params.get('targetLoadType') or
        params.get('loadTargetLoadType') or
        ''
    ).strip().lower()
    result['target_load_type'] = target_load_type
    
    # Target Database Type (to detect Hive)
    target_db_type = (
        params.get('loadTargetType') or
        params.get('load_target_type') or
        params.get('target_type') or
        params.get('targetType') or
        ''
    ).strip()
    result['target_db_type'] = target_db_type
    
    # Hive-specific fields
    if target_db_type.lower() == 'hive':
        # Hadoop Path (required for Hive)
        hadoop_path = (
            params.get('hadoop_path') or
            params.get('hadoopPath') or
            params.get('loadTargetHadoopPath') or
            ''
        ).strip()
        result['hadoop_path'] = hadoop_path
        
        # Partition ID (defaults to N/A)
        partition_id = (
            params.get('partition_id') or
            params.get('partitionId') or
            params.get('loadTargetPartitionId') or
            ''
        ).strip()
        result['partition_id'] = partition_id if partition_id else 'N/A'
    else:
        result['hadoop_path'] = ''
        result['partition_id'] = ''
    
    return result


def validate_data_load_params(params):
    """
    Validate Data Load specific parameters.
    
    Returns (is_valid, error_message) tuple.
    """
    load_fields = extract_data_load_fields(params)
    
    # Validate target_load_type is present
    if not load_fields['target_load_type']:
        return (False, 'Target Load Type is required (overwrite or append)')
    
    if load_fields['target_load_type'] not in ['overwrite', 'append']:
        return (False, f'Invalid Target Load Type: {load_fields["target_load_type"]}. Must be "overwrite" or "append"')
    
    # If target is Hive, require hadoop_path
    if load_fields['target_db_type'].lower() == 'hive':
        if not load_fields['hadoop_path']:
            return (False, 'Hadoop Path is required when target database is Hive')
    
    return (True, None)


def extract_file_load_source_fields(params):
    """
    Extract File Load SOURCE fields (file-based) from request params.
    
    Returns dict with:
    - source_file_type: CSV, JSON, XML, Parquet, Excel
    - source_location_type: unix, hadoop, upload
    - source_unix_path / source_hadoop_path
    - source_delimiter
    - source_file_sql_query
    - source_key_columns
    """
    result = {}
    
    # File Type
    source_file_type = (
        params.get('source_file_type') or
        params.get('sourceFileType') or
        params.get('fileloadSourceFileType') or
        ''
    ).strip().lower()
    result['source_file_type'] = source_file_type
    
    # Location Type
    source_location_type = (
        params.get('source_location_type') or
        params.get('sourceLocationType') or
        params.get('fileloadSourceLocationType') or
        ''
    ).strip().lower()
    result['source_location_type'] = source_location_type
    
    # File paths based on location type
    if source_location_type == 'unix':
        result['source_unix_path'] = (
            params.get('source_unix_path') or
            params.get('sourceUnixPath') or
            params.get('fileloadSourceUnixPath') or
            ''
        ).strip()
        result['source_hadoop_path'] = ''
    elif source_location_type == 'hadoop':
        result['source_hadoop_path'] = (
            params.get('source_hadoop_path') or
            params.get('sourceHadoopPath') or
            params.get('fileloadSourceHadoopPath') or
            ''
        ).strip()
        result['source_unix_path'] = ''
    else:
        result['source_unix_path'] = ''
        result['source_hadoop_path'] = ''
    
    # Delimiter (for CSV)
    source_delimiter = (
        params.get('source_delimiter') or
        params.get('sourceDelimiter') or
        params.get('fileloadSourceDelimiter') or
        ','
    ).strip()
    result['source_delimiter'] = source_delimiter
    
    # SQL Query (default to N/A)
    source_sql = (
        params.get('source_file_sql_query') or
        params.get('sourceFileSqlQuery') or
        params.get('fileloadSourceFileSqlQuery') or
        ''
    ).strip()
    result['source_file_sql_query'] = source_sql if source_sql else 'N/A'
    
    # Key Columns
    source_key_columns = (
        params.get('source_key_columns') or
        params.get('sourceKeyColumns') or
        params.get('fileloadSourceKeyColumns') or
        ''
    ).strip()
    result['source_key_columns'] = source_key_columns
    
    return result


def extract_file_load_target_fields(params):
    """
    Extract File Load TARGET fields (database) from request params.
    Same as Data Load target fields.
    
    Returns dict with:
    - target_load_type: 'overwrite' or 'append'
    - target_db_type: Database type (for Hive detection)
    - hadoop_path: Hadoop path (for Hive)
    - partition_id: Partition ID (for Hive, defaults to N/A)
    """
    result = {}
    
    # Target Load Type (overwrite/append)
    target_load_type = (
        params.get('target_load_type') or
        params.get('targetLoadType') or
        params.get('fileloadTargetLoadType') or
        ''
    ).strip().lower()
    result['target_load_type'] = target_load_type
    
    # Target Database Type (to detect Hive)
    target_db_type = (
        params.get('fileloadTargetType') or
        params.get('fileload_target_type') or
        params.get('target_type') or
        params.get('targetType') or
        ''
    ).strip()
    result['target_db_type'] = target_db_type
    
    # Hive-specific fields
    if target_db_type.lower() == 'hive':
        # Hadoop Path (required for Hive)
        hadoop_path = (
            params.get('hadoop_path') or
            params.get('hadoopPath') or
            params.get('fileloadTargetHadoopPath') or
            ''
        ).strip()
        result['hadoop_path'] = hadoop_path
        
        # Partition ID (defaults to N/A)
        partition_id = (
            params.get('partition_id') or
            params.get('partitionId') or
            params.get('fileloadTargetPartitionId') or
            ''
        ).strip()
        result['partition_id'] = partition_id if partition_id else 'N/A'
    else:
        result['hadoop_path'] = ''
        result['partition_id'] = ''
    
    return result


def validate_file_load_params(params):
    """
    Validate File Load specific parameters (File-to-Database).
    
    Returns (is_valid, error_message) tuple.
    """
    source_fields = extract_file_load_source_fields(params)
    target_fields = extract_file_load_target_fields(params)
    
    # Validate source file type is present
    if not source_fields['source_file_type']:
        return (False, 'Source File Type is required')
    
    # Validate source location type is present
    if not source_fields['source_location_type']:
        return (False, 'Source Location Type is required')
    
    # Validate path based on location type
    if source_fields['source_location_type'] == 'unix' and not source_fields['source_unix_path']:
        return (False, 'Source Unix Path is required when location type is Unix')
    
    if source_fields['source_location_type'] == 'hadoop' and not source_fields['source_hadoop_path']:
        return (False, 'Source Hadoop Path is required when location type is Hadoop')
    
    # Validate target_load_type is present
    if not target_fields['target_load_type']:
        return (False, 'Target Load Type is required (overwrite or append)')
    
    if target_fields['target_load_type'] not in ['overwrite', 'append']:
        return (False, f'Invalid Target Load Type: {target_fields["target_load_type"]}. Must be "overwrite" or "append"')
    
    # If target is Hive, require hadoop_path
    if target_fields['target_db_type'].lower() == 'hive':
        if not target_fields['hadoop_path']:
            return (False, 'Hadoop Path is required when target database is Hive')
    
    return (True, None)

# ============================================================
# UNIX SSH CONFIG FROM ENVIRONMENT VARIABLES
# ============================================================
def get_unix_config():
    """Get Unix server configuration from environment variables"""
    return {
        'host': os.environ.get('UNIX_HOST', 'localhost'),
        'port': int(os.environ.get('UNIX_PORT', '22')),
        'username': os.environ.get('UNIX_USER', 'user'),
        'auth_method': os.environ.get('UNIX_AUTH_METHOD', 'password'),  # 'key' or 'password'
        'private_key': os.environ.get('UNIX_PRIVATE_KEY', ''),
        'password': os.environ.get('UNIX_PASSWORD', '')
    }

def get_unix_config_from_params(params):
    """
    Get Unix server configuration from request params with environment variable fallback.
    
    UI-provided unix_host MUST override any default or env-based host.
    Returns error dict if unix_host is missing.
    
    Expected field names (checked in order):
    - unix_host / unixHost / unixHostname / executionHost
    - user_email / userEmail (for audit/logging purposes)
    - unix_username / unixUsername / unix_user / unixUser
    - unix_password / unixPassword
    - batch_id / batchId
    
    NOTE: SSH port is ALWAYS hardcoded to 22.
    NOTE: SSH key authentication has been removed. Password auth only.
    """
    # Extract Unix host - check multiple possible field names
    unix_host = (
        params.get('unix_host') or 
        params.get('unixHost') or 
        params.get('unixHostname') or 
        params.get('executionHost') or
        params.get('comparisonUnixHost') or
        params.get('loadUnixHost') or
        params.get('schemaUnixHost') or
        params.get('fileloadUnixHost') or
        params.get('filedownloadUnixHost') or
        params.get('mismatchUnixHost') or
        ''
    ).strip()
    
    # Validate unix_host is provided - do not default to localhost
    if not unix_host:
        return {
            'error': 'Unix host is required for execution. Please provide a valid Unix host in the form.'
        }
    
    # SSH port is ALWAYS 22 - hardcoded, not from UI
    unix_port = 22
    
    # Extract User Email (for audit/logging purposes)
    user_email = (
        params.get('user_email') or
        params.get('userEmail') or
        params.get('comparisonUserEmail') or
        params.get('loadUserEmail') or
        params.get('schemaUserEmail') or
        params.get('fileloadUserEmail') or
        params.get('filedownloadUserEmail') or
        params.get('mismatchUserEmail') or
        ''
    ).strip()
    
    # Extract Unix username
    unix_username = (
        params.get('unix_username') or 
        params.get('unixUsername') or
        params.get('unix_user') or
        params.get('unixUser') or
        params.get('comparisonUnixUsername') or
        params.get('loadUnixUsername') or
        params.get('schemaUnixUsername') or
        params.get('fileloadUnixUsername') or
        params.get('filedownloadUnixUsername') or
        params.get('mismatchUnixUsername') or
        os.environ.get('UNIX_USER', 'user')
    ).strip()
    
    # Extract Unix password - do NOT strip internal spaces, only ends
    unix_password = (
        params.get('unix_password') or 
        params.get('unixPassword') or
        params.get('comparisonUnixPassword') or
        params.get('loadUnixPassword') or
        params.get('schemaUnixPassword') or
        params.get('fileloadUnixPassword') or
        params.get('filedownloadUnixPassword') or
        params.get('mismatchUnixPassword') or
        os.environ.get('UNIX_PASSWORD', '')
    )
    if unix_password:
        unix_password = unix_password.lstrip().rstrip()  # Safe strip
    
    # Extract Batch ID
    batch_id = (
        params.get('batch_id') or 
        params.get('batchId') or
        params.get('comparisonBatchId') or
        params.get('loadBatchId') or
        params.get('schemaBatchId') or
        params.get('fileloadBatchId') or
        params.get('filedownloadBatchId') or
        params.get('mismatchBatchId') or
        ''
    ).strip()
    
    # Always use password authentication (SSH key auth removed)
    return {
        'host': unix_host,
        'port': unix_port,  # Always 22
        'username': unix_username,
        'auth_method': 'password',
        'password': unix_password,
        'batch_id': batch_id,
        'user_email': user_email  # Captured for audit/logging
    }


def extract_file_path_fields(params):
    """
    Extract file path fields from request params.
    
    Handles location_type (unix/hadoop) and returns the appropriate path fields.
    Also handles file_sql_query - defaults to "N/A" if empty.
    
    Returns dict with:
    - source_unix_path / source_hadoop_path
    - target_unix_path / target_hadoop_path
    - source_file_sql_query / target_file_sql_query
    - source_location_type / target_location_type
    """
    result = {}
    
    # Source location type handling
    source_location_type = (
        params.get('source_location_type') or
        params.get('sourceLocationType') or
        params.get('comparisonSourceLocationType') or
        ''
    ).strip().lower()
    
    result['source_location_type'] = source_location_type
    
    if source_location_type == 'unix':
        result['source_unix_path'] = (
            params.get('source_unix_path') or
            params.get('sourceUnixPath') or
            params.get('comparisonSourceUnixPath') or
            ''
        ).strip()
        result['source_hadoop_path'] = ''
    elif source_location_type == 'hadoop':
        result['source_hadoop_path'] = (
            params.get('source_hadoop_path') or
            params.get('sourceHadoopPath') or
            params.get('comparisonSourceHadoopPath') or
            ''
        ).strip()
        result['source_unix_path'] = ''
    else:
        # Fallback: check for generic file path
        result['source_unix_path'] = (
            params.get('source_file_path') or
            params.get('sourceFilePath') or
            params.get('comparisonSourceFilePath') or
            ''
        ).strip()
        result['source_hadoop_path'] = ''
    
    # Target location type handling
    target_location_type = (
        params.get('target_location_type') or
        params.get('targetLocationType') or
        params.get('comparisonTargetLocationType') or
        ''
    ).strip().lower()
    
    result['target_location_type'] = target_location_type
    
    if target_location_type == 'unix':
        result['target_unix_path'] = (
            params.get('target_unix_path') or
            params.get('targetUnixPath') or
            params.get('comparisonTargetUnixPath') or
            ''
        ).strip()
        result['target_hadoop_path'] = ''
    elif target_location_type == 'hadoop':
        result['target_hadoop_path'] = (
            params.get('target_hadoop_path') or
            params.get('targetHadoopPath') or
            params.get('comparisonTargetHadoopPath') or
            ''
        ).strip()
        result['target_unix_path'] = ''
    else:
        # Fallback: check for generic file path
        result['target_unix_path'] = (
            params.get('target_file_path') or
            params.get('targetFilePath') or
            params.get('comparisonTargetFilePath') or
            ''
        ).strip()
        result['target_hadoop_path'] = ''
    
    # Handle file SQL query - default to "N/A" if empty
    source_file_sql = (
        params.get('source_file_sql_query') or
        params.get('sourceFileSqlQuery') or
        params.get('comparisonSourceFileSqlQuery') or
        ''
    ).strip()
    result['source_file_sql_query'] = source_file_sql if source_file_sql else 'N/A'
    
    target_file_sql = (
        params.get('target_file_sql_query') or
        params.get('targetFileSqlQuery') or
        params.get('comparisonTargetFileSqlQuery') or
        ''
    ).strip()
    result['target_file_sql_query'] = target_file_sql if target_file_sql else 'N/A'
    
    return result


# ============================================================
# ROUTES
# ============================================================
@app.route('/')
def index():
    return render_template('index.html')

# ============================================================
# STATIC EXECUTION API - POST /api/static/run
# ============================================================
@app.route('/api/static/run', methods=['POST'])
def static_run():
    """
    Generic execution endpoint for static UI.
    Request payload:
    {
        "tab": "data_comparison" | "data_load" | "schema_generation" | "file_load" | "file_download",
        "params": { all form fields visible in the selected tab }
    }
    Returns: { "run_id": "uuid", "status": "PENDING" }
    """
    try:
        data = request.json
        tab = data.get('tab')
        params = data.get('params', {})
        
        if not tab or tab not in TAB_SCRIPT_MAP:
            return jsonify({
                'error': f'Invalid tab: {tab}. Must be one of: {list(TAB_SCRIPT_MAP.keys())}'
            }), 400
        
        # Generate unique run ID
        run_id = str(uuid.uuid4())[:8]
        
        # Initialize run entry
        run_registry[run_id] = {
            'status': 'PENDING',
            'logs': [],
            'tab': tab,
            'params': params,
            'thread': None,
            'ssh_client': None,
            'stop_requested': False
        }
        
        # Start execution in background thread
        thread = threading.Thread(
            target=execute_static_operation,
            args=(run_id, tab, params)
        )
        thread.daemon = True
        run_registry[run_id]['thread'] = thread
        thread.start()
        
        return jsonify({
            'run_id': run_id,
            'status': 'PENDING',
            'message': f'{tab} operation started'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================
# STATIC LOGS API - GET /api/static/logs/<run_id> (SSE)
# ============================================================
@app.route('/api/static/logs/<run_id>')
def static_logs(run_id):
    """
    Server-Sent Events endpoint for real-time log streaming.
    Streams stdout + stderr line by line until completion.
    """
    def generate():
        if run_id not in run_registry:
            yield f"data: {json.dumps({'type': 'error', 'message': 'Run ID not found'})}\n\n"
            return
        
        run_entry = run_registry[run_id]
        last_index = 0
        
        # Send initial status
        yield f"data: {json.dumps({'type': 'status', 'status': run_entry['status']})}\n\n"
        
        while True:
            # Check if run exists
            if run_id not in run_registry:
                yield f"data: {json.dumps({'type': 'end', 'message': 'Run cleared'})}\n\n"
                break
            
            run_entry = run_registry[run_id]
            
            # Send any new logs
            current_logs = run_entry['logs']
            if len(current_logs) > last_index:
                for log in current_logs[last_index:]:
                    yield f"data: {json.dumps({'type': 'log', 'message': log})}\n\n"
                last_index = len(current_logs)
            
            # Check if execution is complete
            status = run_entry['status']
            if status in ['SUCCESS', 'FAILED', 'STOPPED', 'PARTIAL_SUCCESS']:
                # For multi-table runs, include summary
                summary = {}
                if run_entry.get('rows_status'):
                    rows_status = run_entry['rows_status']
                    summary = {
                        'total_rows': run_entry.get('total_rows', 0),
                        'success_count': sum(1 for r in rows_status if r.get('status') == 'SUCCESS'),
                        'failed_count': sum(1 for r in rows_status if r.get('status') == 'FAILED')
                    }
                
                # For single_table data_comparison, include download info
                download_info = {}
                if run_entry.get('download_available'):
                    download_info = {
                        'download_available': True,
                        'download_url': f'/api/static/comparison/download/{run_id}',
                        'download_filename': run_entry.get('download_filename', 'output.csv')
                    }
                
                yield f"data: {json.dumps({'type': 'status', 'status': status, 'summary': summary, **download_info})}\n\n"
                yield f"data: {json.dumps({'type': 'end', 'status': status, 'summary': summary, **download_info})}\n\n"
                break
            
            # Small delay to prevent CPU spinning
            time.sleep(0.1)
    
    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no'
        }
    )

# ============================================================
# CONTROL ENDPOINTS
# ============================================================
@app.route('/api/static/stop/<run_id>', methods=['POST'])
def static_stop(run_id):
    """Stop execution for a given run_id"""
    if run_id not in run_registry:
        return jsonify({'error': 'Run ID not found'}), 404
    
    run_entry = run_registry[run_id]
    run_entry['stop_requested'] = True
    
    # Close SSH connection if active
    if run_entry.get('ssh_client'):
        try:
            run_entry['ssh_client'].close()
        except:
            pass
    
    run_entry['status'] = 'STOPPED'
    add_log(run_id, '⛔ Execution stopped by user')
    
    return jsonify({'status': 'STOPPED', 'message': 'Execution stopped'})

@app.route('/api/static/clear/<run_id>', methods=['POST'])
def static_clear(run_id):
    """Clear logs for a given run_id"""
    if run_id not in run_registry:
        return jsonify({'error': 'Run ID not found'}), 404
    
    run_registry[run_id]['logs'] = []
    
    return jsonify({'status': 'cleared', 'message': 'Logs cleared'})

@app.route('/api/static/status/<run_id>')
def static_status(run_id):
    """Get current status for a run"""
    if run_id not in run_registry:
        return jsonify({'error': 'Run ID not found'}), 404
    
    run_entry = run_registry[run_id]
    response = {
        'run_id': run_id,
        'status': run_entry['status'],
        'log_count': len(run_entry['logs']),
        'tab': run_entry['tab']
    }
    
    # Include download info if available
    if run_entry.get('download_available'):
        response['download_available'] = True
        response['download_filename'] = run_entry.get('download_filename', 'output.csv')
    
    return jsonify(response)


# ============================================================
# SINGLE TABLE COMPARISON - FILE DOWNLOAD ENDPOINT
# ============================================================
@app.route('/api/static/comparison/download/<run_id>')
def static_comparison_download(run_id):
    """
    Download the comparison output file from Unix server /tmp/ folder.
    
    This endpoint is for single_table mode only.
    After the comparison script completes, this fetches the output CSV
    from the Unix server and returns it as a downloadable file.
    
    Output filename rules:
    - database_to_database → {target_table}.csv
    - file_to_database → {target_table}.csv
    - file_to_file → file_to_file.csv
    - database_to_file → db_to_file.csv
    """
    try:
        if run_id not in run_registry:
            return jsonify({'error': 'Run ID not found'}), 404
        
        run_entry = run_registry[run_id]
        
        # Validate this is a completed single_table data_comparison run
        if run_entry['tab'] != 'data_comparison':
            return jsonify({'error': 'Download is only available for data comparison runs'}), 400
        
        if run_entry['status'] not in ['SUCCESS', 'PARTIAL_SUCCESS']:
            return jsonify({'error': f'Cannot download - run status is {run_entry["status"]}'}), 400
        
        # Get download info from run_entry
        if not run_entry.get('download_available'):
            return jsonify({'error': 'No download available for this run. This may be a multiple_tables run.'}), 400
        
        unix_source_path = run_entry.get('unix_output_path')
        download_filename = run_entry.get('download_filename', 'output.csv')
        
        if not unix_source_path:
            return jsonify({'error': 'Output file path not found in run registry'}), 500
        
        # Get Unix config from params
        params = run_entry.get('params', {})
        unix_config = get_unix_config_from_params(params)
        
        if 'error' in unix_config:
            return jsonify({'error': unix_config['error']}), 400
        
        print(f"⬇️ Preparing download from Unix: {unix_source_path} as {download_filename}")
        
        # Connect to Unix server via SFTP
        try:
            ssh_client = paramiko.SSHClient()
            ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            
            ssh_client.connect(
                hostname=unix_config['host'],
                port=unix_config['port'],
                username=unix_config['username'],
                password=unix_config['password'],
                timeout=30
            )
            
            sftp = ssh_client.open_sftp()
            
            # Create a temporary file to store the downloaded content
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
            temp_path = temp_file.name
            temp_file.close()
            
            try:
                # Download file from Unix server
                sftp.get(unix_source_path, temp_path)
                print(f"✅ Downloaded file from {unix_source_path} to {temp_path}")
            except FileNotFoundError:
                # Try alternative paths in /tmp
                alt_paths = [
                    '/tmp/output.csv',
                    '/tmp/comparison_output.csv',
                    f'/tmp/{download_filename}'
                ]
                
                file_found = False
                for alt_path in alt_paths:
                    try:
                        sftp.get(alt_path, temp_path)
                        print(f"✅ Found and downloaded file from alternative path: {alt_path}")
                        file_found = True
                        break
                    except FileNotFoundError:
                        continue
                
                if not file_found:
                    # List files in /tmp and find the most recent CSV
                    try:
                        tmp_files = sftp.listdir_attr('/tmp')
                        csv_files = [f for f in tmp_files if f.filename.endswith('.csv')]
                        
                        if csv_files:
                            # Sort by modification time (most recent first)
                            csv_files.sort(key=lambda x: x.st_mtime, reverse=True)
                            most_recent = csv_files[0].filename
                            sftp.get(f'/tmp/{most_recent}', temp_path)
                            print(f"✅ Downloaded most recent CSV from /tmp: {most_recent}")
                            file_found = True
                    except Exception as e:
                        print(f"❌ Error listing /tmp directory: {str(e)}")
                
                if not file_found:
                    sftp.close()
                    ssh_client.close()
                    os.unlink(temp_path)
                    print(f"❌ Output CSV not found in /tmp for this run")
                    return jsonify({'error': 'Output CSV not found in /tmp for this run'}), 404
            
            sftp.close()
            ssh_client.close()
            
            # Return the file as a download
            return send_file(
                temp_path,
                mimetype='text/csv',
                as_attachment=True,
                download_name=download_filename
            )
            
        except Exception as e:
            print(f"❌ SFTP error: {str(e)}")
            return jsonify({'error': f'Failed to download file from Unix server: {str(e)}'}), 500
            
    except Exception as e:
        print(f"❌ Download error: {str(e)}")
        return jsonify({'error': str(e)}), 500


# ============================================================
# FILE DOWNLOAD TAB - VALIDATION AND DOWNLOAD ENDPOINT
# ============================================================
def extract_file_download_fields(params):
    """
    Extract File Download specific fields from request params.
    
    Returns dict with:
    - source_type: 'file' or 'database'
    - target_file_type: CSV, TSV, JSON, Parquet, XML
    - target_delimiter: Delimiter for CSV/TSV
    - target_file_name: Output file name (mandatory)
    - number_of_rows: Max rows to download (max 10000)
    """
    result = {}
    
    # Source type
    source_type = (
        params.get('source_type') or
        params.get('sourceType') or
        ''
    ).strip().lower()
    result['source_type'] = source_type
    
    # Target file type
    target_file_type = (
        params.get('target_file_type') or
        params.get('targetFileType') or
        params.get('filedownloadTargetFileType') or
        ''
    ).strip()
    result['target_file_type'] = target_file_type
    
    # Target delimiter (for CSV/TSV)
    target_delimiter = (
        params.get('target_delimiter') or
        params.get('targetDelimiter') or
        params.get('filedownloadTargetDelimiter') or
        ','
    ).strip()
    result['target_delimiter'] = target_delimiter
    
    # Target file name (mandatory)
    target_file_name = (
        params.get('target_file_name') or
        params.get('targetFileName') or
        params.get('filedownloadTargetFileName') or
        ''
    ).strip()
    result['target_file_name'] = target_file_name
    
    # Number of rows (max 100000 for File Download tab)
    number_of_rows_str = (
        params.get('number_of_rows') or
        params.get('numberOfRows') or
        params.get('filedownloadTargetNumberOfRows') or
        ''
    )
    
    if number_of_rows_str:
        try:
            number_of_rows = int(number_of_rows_str)
            result['number_of_rows'] = number_of_rows
        except ValueError:
            result['number_of_rows'] = None
    else:
        result['number_of_rows'] = None
    
    return result


def validate_file_download_params(params):
    """
    Validate File Download specific parameters.
    
    Returns (is_valid, error_message) tuple.
    """
    fields = extract_file_download_fields(params)
    
    # Validate source_type is present
    if not fields['source_type']:
        return (False, 'Source Type is required (file or database)')
    
    if fields['source_type'] not in ['file', 'database']:
        return (False, f'Invalid Source Type: {fields["source_type"]}. Must be "file" or "database"')
    
    # Validate target_file_name is present (mandatory)
    if not fields['target_file_name']:
        return (False, 'Target File Name is required')
    
    # Validate number_of_rows does not exceed 100000 (File Download tab limit)
    if fields['number_of_rows'] is not None:
        if fields['number_of_rows'] > 100000:
            return (False, 'Number of Rows cannot exceed 100,000')
        if fields['number_of_rows'] < 1:
            return (False, 'Number of Rows must be at least 1')
    
    # Validate target_file_type
    if not fields['target_file_type']:
        return (False, 'Target File Type is required')
    
    valid_file_types = ['CSV', 'TSV', 'JSON', 'Parquet', 'XML', 'Excel', 'csv', 'tsv', 'json', 'parquet', 'xml', 'excel', 'XLSX', 'xlsx']
    if fields['target_file_type'] not in valid_file_types:
        return (False, f'Invalid Target File Type: {fields["target_file_type"]}')
    
    # Validate delimiter is present for CSV/TSV
    if fields['target_file_type'].upper() in ['CSV', 'TSV']:
        if not fields['target_delimiter']:
            return (False, 'Delimiter is required for CSV/TSV file types')
    
    return (True, None)


# ============================================================
# FILE DOWNLOAD TAB - WITH LOCAL FILE UPLOAD ENDPOINT
# ============================================================
@app.route('/api/static/filedownload/run-with-upload', methods=['POST'])
def static_filedownload_run_with_upload():
    """
    Execute File Download operation with local file upload.
    
    This endpoint handles multipart/form-data submissions where:
    - source_local_file: The uploaded file from user's local system
    - The file is transferred to Unix /tmp/{batch_id}_{original_filename}
    - The Unix path is then used as source input for the execution script
    
    Returns: { "run_id": "uuid", "status": "PENDING" }
    """
    try:
        # Debug logging
        print(f"[FileDownload Upload] Files received: {list(request.files.keys())}")
        print(f"[FileDownload Upload] Form fields received: {list(request.form.keys())}")
        
        # Check for uploaded file
        if 'source_local_file' not in request.files:
            return jsonify({
                'error': 'No file uploaded. Please select a local file to upload.'
            }), 400
        
        uploaded_file = request.files['source_local_file']
        
        if uploaded_file.filename == '':
            return jsonify({
                'error': 'No file selected. Please select a local file to upload.'
            }), 400
        
        # Read file content into memory
        file_content = uploaded_file.read()
        original_filename = uploaded_file.filename
        
        print(f"[FileDownload Upload] File received: {original_filename}, size: {len(file_content)} bytes")
        
        # Collect form parameters
        params = {}
        for key in request.form:
            params[key] = request.form[key]
        
        # Validate required fields
        target_file_name = params.get('target_file_name', '').strip()
        if not target_file_name:
            return jsonify({'error': 'Target File Name is required'}), 400
        
        # Validate number_of_rows (max 100000)
        number_of_rows_str = params.get('number_of_rows', '')
        if number_of_rows_str:
            try:
                number_of_rows = int(number_of_rows_str)
                if number_of_rows > 100000:
                    return jsonify({'error': 'Number of Rows cannot exceed 100,000'}), 400
                if number_of_rows < 1:
                    return jsonify({'error': 'Number of Rows must be at least 1'}), 400
            except ValueError:
                return jsonify({'error': 'Number of Rows must be a valid number'}), 400
        
        # Get Unix config from params
        unix_config = get_unix_config_from_params(params)
        if 'error' in unix_config:
            return jsonify({'error': unix_config['error']}), 400
        
        # Resolve execution context
        exec_context = resolve_execution_context(params)
        if exec_context['error']:
            return jsonify({'error': exec_context['error']}), 400
        
        batch_id = exec_context['batch_id']
        
        # Generate unique run ID
        run_id = str(uuid.uuid4())[:8]
        
        # Store file info for background thread
        # Preserve original filename (no batch_id prefix) - just sanitize special characters
        safe_filename = original_filename.replace(' ', '_').replace('/', '_').replace('\\', '_')
        unix_dest_path = f'/tmp/{safe_filename}'
        
        # Initialize run entry
        run_registry[run_id] = {
            'status': 'PENDING',
            'logs': [],
            'tab': 'file_download',
            'params': params,
            'thread': None,
            'ssh_client': None,
            'stop_requested': False,
            'uploaded_file_content': file_content,
            'uploaded_file_name': original_filename,
            'unix_file_path': unix_dest_path
        }
        
        # Start execution in background thread
        thread = threading.Thread(
            target=execute_filedownload_with_upload,
            args=(run_id, params, file_content, original_filename, unix_dest_path)
        )
        thread.daemon = True
        run_registry[run_id]['thread'] = thread
        thread.start()
        
        return jsonify({
            'run_id': run_id,
            'status': 'PENDING',
            'message': f'File Download operation started with uploaded file: {original_filename}'
        })
        
    except Exception as e:
        print(f"[FileDownload Upload] Error: {str(e)}")
        return jsonify({'error': str(e)}), 500


def execute_filedownload_with_upload(run_id, params, file_content, original_filename, unix_dest_path):
    """
    Execute File Download operation with local file upload.
    
    Steps:
    1. Connect to Unix server via SSH/SFTP
    2. Upload the file to /tmp/{original_filename} (preserving original filename)
    3. Update params with the Unix file path
    4. Execute the File Download script
    """
    try:
        run_registry[run_id]['status'] = 'RUNNING'
        add_log(run_id, f'🚀 Starting File Download with local file upload...')
        add_log(run_id, f'📤 Uploaded file: {original_filename} ({len(file_content)} bytes)')
        
        # Resolve execution context
        exec_context = resolve_execution_context(params, run_id)
        
        if exec_context['error']:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, f'❌ {exec_context["error"]}')
            return
        
        # Extract resolved context
        working_dir = exec_context['working_dir']
        batch_id = exec_context['batch_id']
        env = exec_context['env']
        unix_config = exec_context['unix_config']
        
        # Store in run registry
        set_execution_context(run_id, working_dir, batch_id, env)
        
        add_log(run_id, f'📂 Working directory: {working_dir}')
        add_log(run_id, f'🌍 Environment: {env}, Batch ID: {batch_id}')
        add_log(run_id, f'📡 Connecting to Unix server {unix_config["host"]}:{unix_config["port"]}...')
        
        # Connect to Unix server
        ssh_client = connect_unix_server(unix_config, run_id)
        
        if not ssh_client:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, '❌ Failed to connect to Unix server')
            return
        
        run_registry[run_id]['ssh_client'] = ssh_client
        add_log(run_id, '✅ Connected to Unix server')
        
        # Step 1: Upload file to Unix /tmp via SFTP
        add_log(run_id, f'📤 Uploading file to Unix: {unix_dest_path}')
        
        try:
            sftp = ssh_client.open_sftp()
            
            # Write file content to Unix
            with sftp.file(unix_dest_path, 'wb') as remote_file:
                remote_file.write(file_content)
            
            sftp.close()
            add_log(run_id, f'✅ File uploaded successfully to {unix_dest_path}')
            
        except Exception as e:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, f'❌ Failed to upload file to Unix: {str(e)}')
            ssh_client.close()
            return
        
        # Step 2: Update params with Unix file path
        params['filedownloadSourceUnixPath'] = unix_dest_path
        params['filedownloadSourceLocationType'] = 'unix'  # Override to unix since file is now on Unix
        
        # Check if stop requested
        if run_registry[run_id]['stop_requested']:
            ssh_client.close()
            return
        
        # Step 3: Execute the File Download script
        script_path = TAB_SCRIPT_MAP['file_download']
        add_log(run_id, f'📋 Using script: {script_path}')
        
        # Ensure active_tab is in params
        params['active_tab'] = 'file_download'
        add_log(run_id, f'🏷️ Active tab passed to Unix: file_download')
        
        # Build script arguments
        script_args = build_script_arguments(params)
        
        # Build full command
        add_log(run_id, f'▶️ Executing script: {script_path}')
        command = f'cd {shlex.quote(working_dir)} && {script_path} {script_args}'
        
        # Execute command and stream output
        success = execute_command_streaming(ssh_client, command, run_id)
        
        # Close SSH connection
        ssh_client.close()
        run_registry[run_id]['ssh_client'] = None
        add_log(run_id, '🔌 Unix server connection closed')
        
        # Set final status
        if run_registry[run_id]['stop_requested']:
            run_registry[run_id]['status'] = 'STOPPED'
        elif success:
            run_registry[run_id]['status'] = 'SUCCESS'
            add_log(run_id, f'✅ File Download completed successfully!')
            
            # Set up download info
            target_file_name = params.get('target_file_name', 'output.csv')
            run_registry[run_id]['download_available'] = True
            run_registry[run_id]['download_filename'] = target_file_name
            run_registry[run_id]['unix_output_path'] = '/tmp/output.csv'
            add_log(run_id, f'⬇️ Output file ready for download: {target_file_name}')
        else:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, f'❌ File Download failed')
            
    except Exception as e:
        run_registry[run_id]['status'] = 'FAILED'
        add_log(run_id, f'❌ Error: {str(e)}')
        print(f"Error in execute_filedownload_with_upload: {str(e)}")


# ============================================================
# FILE LOAD TAB - WITH LOCAL FILE UPLOAD ENDPOINT
# ============================================================
@app.route('/api/static/fileload/run-with-upload', methods=['POST'])
def static_fileload_run_with_upload():
    """
    Execute File Load operation with local file upload.
    
    This endpoint handles multipart/form-data submissions where:
    - source_local_file: The uploaded file from user's local system
    - The file is transferred to Unix /tmp/{original_filename} (preserving original name)
    - The Unix path is then used as source input for the execution script
    
    Returns: { "run_id": "uuid", "status": "PENDING" }
    """
    try:
        # Debug logging
        print(f"[FileLoad Upload] Files received: {list(request.files.keys())}")
        print(f"[FileLoad Upload] Form fields received: {list(request.form.keys())}")
        
        # Check for uploaded file
        if 'source_local_file' not in request.files:
            return jsonify({
                'error': 'No file uploaded. Please select a local file to upload.'
            }), 400
        
        uploaded_file = request.files['source_local_file']
        
        if uploaded_file.filename == '':
            return jsonify({
                'error': 'No file selected. Please select a local file to upload.'
            }), 400
        
        # Read file content into memory
        file_content = uploaded_file.read()
        original_filename = uploaded_file.filename
        
        print(f"[FileLoad Upload] File received: {original_filename}, size: {len(file_content)} bytes")
        
        # Collect form parameters
        params = {}
        for key in request.form:
            params[key] = request.form[key]
        
        # Validate File Load specific params
        is_valid, error_msg = validate_file_load_params_for_upload(params)
        if not is_valid:
            return jsonify({'error': error_msg}), 400
        
        # Get Unix config from params
        unix_config = get_unix_config_from_params(params)
        if 'error' in unix_config:
            return jsonify({'error': unix_config['error']}), 400
        
        # Resolve execution context
        exec_context = resolve_execution_context(params)
        if exec_context['error']:
            return jsonify({'error': exec_context['error']}), 400
        
        batch_id = exec_context['batch_id']
        
        # Generate unique run ID
        run_id = str(uuid.uuid4())[:8]
        
        # Store file info for background thread
        # Preserve original filename (no batch_id prefix) - just sanitize special characters
        safe_filename = original_filename.replace(' ', '_').replace('/', '_').replace('\\', '_')
        unix_dest_path = f'/tmp/{safe_filename}'
        
        # Initialize run entry
        run_registry[run_id] = {
            'status': 'PENDING',
            'logs': [],
            'tab': 'file_load',
            'params': params,
            'thread': None,
            'ssh_client': None,
            'stop_requested': False,
            'uploaded_file_content': file_content,
            'uploaded_file_name': original_filename,
            'unix_file_path': unix_dest_path
        }
        
        # Start execution in background thread
        thread = threading.Thread(
            target=execute_fileload_with_upload,
            args=(run_id, params, file_content, original_filename, unix_dest_path)
        )
        thread.daemon = True
        run_registry[run_id]['thread'] = thread
        thread.start()
        
        return jsonify({
            'run_id': run_id,
            'status': 'PENDING',
            'message': f'File Load operation started with uploaded file: {original_filename}'
        })
        
    except Exception as e:
        print(f"[FileLoad Upload] Error: {str(e)}")
        return jsonify({'error': str(e)}), 500


def validate_file_load_params_for_upload(params):
    """
    Validate File Load parameters when using local file upload.
    Similar to validate_file_load_params but skips path validation since file is uploaded.
    """
    target_fields = extract_file_load_target_fields(params)
    
    # Source file type validation
    source_file_type = (
        params.get('source_file_type') or
        params.get('sourceFileType') or
        params.get('fileloadSourceFileType') or
        ''
    ).strip()
    
    if not source_file_type:
        return (False, 'Source File Type is required')
    
    # Validate target_load_type is present
    if not target_fields['target_load_type']:
        return (False, 'Target Load Type is required (overwrite or append)')
    
    if target_fields['target_load_type'] not in ['overwrite', 'append']:
        return (False, f'Invalid Target Load Type: {target_fields["target_load_type"]}. Must be "overwrite" or "append"')
    
    # If target is Hive, require hadoop_path
    if target_fields['target_db_type'].lower() == 'hive':
        if not target_fields['hadoop_path']:
            return (False, 'Hadoop Path is required when target database is Hive')
    
    return (True, None)


def execute_fileload_with_upload(run_id, params, file_content, original_filename, unix_dest_path):
    """
    Execute File Load operation with local file upload.
    
    Steps:
    1. Connect to Unix server via SSH/SFTP
    2. Upload the file to /tmp/{original_filename} (preserving original filename)
    3. Update params with the Unix file path
    4. Execute the File Load script
    """
    try:
        run_registry[run_id]['status'] = 'RUNNING'
        add_log(run_id, f'🚀 Starting File Load with local file upload...')
        add_log(run_id, f'📤 Uploaded file: {original_filename} ({len(file_content)} bytes)')
        
        # Resolve execution context
        exec_context = resolve_execution_context(params, run_id)
        
        if exec_context['error']:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, f'❌ {exec_context["error"]}')
            return
        
        # Extract resolved context
        working_dir = exec_context['working_dir']
        batch_id = exec_context['batch_id']
        env = exec_context['env']
        unix_config = exec_context['unix_config']
        
        # Store in run registry
        set_execution_context(run_id, working_dir, batch_id, env)
        
        add_log(run_id, f'📂 Working directory: {working_dir}')
        add_log(run_id, f'🌍 Environment: {env}, Batch ID: {batch_id}')
        add_log(run_id, f'📡 Connecting to Unix server {unix_config["host"]}:{unix_config["port"]}...')
        
        # Connect to Unix server
        ssh_client = connect_unix_server(unix_config, run_id)
        
        if not ssh_client:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, '❌ Failed to connect to Unix server')
            return
        
        run_registry[run_id]['ssh_client'] = ssh_client
        add_log(run_id, '✅ Connected to Unix server')
        
        # Step 1: Upload file to Unix /tmp via SFTP
        add_log(run_id, f'📤 Uploading file to Unix: {unix_dest_path}')
        
        try:
            sftp = ssh_client.open_sftp()
            
            # Write file content to Unix
            with sftp.file(unix_dest_path, 'wb') as remote_file:
                remote_file.write(file_content)
            
            sftp.close()
            add_log(run_id, f'✅ File uploaded successfully to {unix_dest_path}')
            
        except Exception as e:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, f'❌ Failed to upload file to Unix: {str(e)}')
            ssh_client.close()
            return
        
        # Step 2: Update params with Unix file path
        params['fileloadSourceUnixPath'] = unix_dest_path
        params['source_unix_path'] = unix_dest_path
        params['source_location_type'] = 'unix'  # Override to unix since file is now on Unix
        params['fileloadSourceLocationType'] = 'unix'
        
        # Extract and add File Load source fields to params
        source_fields = extract_file_load_source_fields(params)
        params['source_file_type'] = source_fields['source_file_type']
        params['source_delimiter'] = source_fields['source_delimiter']
        params['source_file_sql_query'] = source_fields['source_file_sql_query']
        params['source_key_columns'] = source_fields['source_key_columns']
        
        # Extract and add File Load target fields
        target_fields = extract_file_load_target_fields(params)
        params['target_load_type'] = target_fields['target_load_type']
        
        add_log(run_id, f'📄 Source File Type: {source_fields["source_file_type"]}')
        add_log(run_id, f'📍 Source Location: unix (uploaded to /tmp)')
        add_log(run_id, f'📂 Source Unix Path: {unix_dest_path}')
        add_log(run_id, f'⚡ Target Load Type: {target_fields["target_load_type"]}')
        
        # Add Hive-specific fields if applicable
        if target_fields['target_db_type'].lower() == 'hive':
            params['hadoop_path'] = target_fields['hadoop_path']
            params['partition_id'] = target_fields['partition_id']
            add_log(run_id, f'🐝 Target DB: Hive')
            add_log(run_id, f'📂 Hadoop Path: {target_fields["hadoop_path"]}')
            add_log(run_id, f'🔖 Partition ID: {target_fields["partition_id"]}')
        
        # Extract and log time_zone
        time_zone = extract_time_zone(params)
        params['time_zone'] = time_zone
        add_log(run_id, f'🕐 Time Zone: {time_zone}')
        
        # Check if stop requested
        if run_registry[run_id]['stop_requested']:
            ssh_client.close()
            run_registry[run_id]['status'] = 'STOPPED'
            add_log(run_id, '⛔ Operation stopped by user')
            return
        
        # Step 3: Execute the File Load script
        script_path = TAB_SCRIPT_MAP['file_load']
        add_log(run_id, f'📋 Using script: {script_path}')
        
        # Ensure active_tab is in params
        params['active_tab'] = 'file_load'
        add_log(run_id, f'🏷️ Active tab passed to Unix: file_load')
        
        # Build script arguments
        script_args = build_script_arguments(params)
        
        # Build full command
        add_log(run_id, f'▶️ Executing script: {script_path}')
        command = f'cd {shlex.quote(working_dir)} && {script_path} {script_args}'
        
        # Execute command and stream output
        success = execute_command_streaming(ssh_client, command, run_id)
        
        # Close SSH connection
        ssh_client.close()
        run_registry[run_id]['ssh_client'] = None
        add_log(run_id, '🔌 Unix server connection closed')
        
        # Set final status
        if run_registry[run_id]['stop_requested']:
            run_registry[run_id]['status'] = 'STOPPED'
        elif success:
            run_registry[run_id]['status'] = 'SUCCESS'
            add_log(run_id, f'✅ File Load completed successfully!')
        else:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, f'❌ File Load failed')
            
    except Exception as e:
        run_registry[run_id]['status'] = 'FAILED'
        add_log(run_id, f'❌ Error: {str(e)}')
        print(f"Error in execute_fileload_with_upload: {str(e)}")


@app.route('/api/static/filedownload/download/<run_id>')
def static_filedownload_download(run_id):
    """
    Download the generated file from Unix server for File Download tab.
    
    This endpoint serves the output file generated by the File Download operation.
    The downloaded file name is the target_file_name specified by the user.
    """
    try:
        if run_id not in run_registry:
            return jsonify({'error': 'Run ID not found'}), 404
        
        run_entry = run_registry[run_id]
        
        # Validate this is a completed file_download run
        if run_entry['tab'] != 'file_download':
            return jsonify({'error': 'Download is only available for file download runs'}), 400
        
        if run_entry['status'] not in ['SUCCESS', 'PARTIAL_SUCCESS']:
            return jsonify({'error': f'Cannot download - run status is {run_entry["status"]}'}), 400
        
        # Get download info from run_entry
        if not run_entry.get('download_available'):
            return jsonify({'error': 'No download available for this run'}), 400
        
        unix_source_path = run_entry.get('unix_output_path')
        download_filename = run_entry.get('download_filename', 'output.csv')
        
        if not unix_source_path:
            return jsonify({'error': 'Output file path not found in run registry'}), 500
        
        # Get Unix config from params
        params = run_entry.get('params', {})
        unix_config = get_unix_config_from_params(params)
        
        if 'error' in unix_config:
            return jsonify({'error': unix_config['error']}), 400
        
        print(f"⬇️ Preparing File Download from Unix: {unix_source_path} as {download_filename}")
        
        # Connect to Unix server via SFTP
        try:
            ssh_client = paramiko.SSHClient()
            ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            
            ssh_client.connect(
                hostname=unix_config['host'],
                port=unix_config['port'],
                username=unix_config['username'],
                password=unix_config['password'],
                timeout=30
            )
            
            sftp = ssh_client.open_sftp()
            
            # Create a temporary file to store the downloaded content
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.tmp')
            temp_path = temp_file.name
            temp_file.close()
            
            try:
                # Download file from Unix server
                sftp.get(unix_source_path, temp_path)
                print(f"✅ Downloaded file from {unix_source_path} to {temp_path}")
            except FileNotFoundError:
                # Try alternative paths in /tmp
                alt_paths = [
                    '/tmp/output.csv',
                    '/tmp/filedownload_output.csv',
                    f'/tmp/{download_filename}'
                ]
                
                file_found = False
                for alt_path in alt_paths:
                    try:
                        sftp.get(alt_path, temp_path)
                        print(f"✅ Found and downloaded file from alternative path: {alt_path}")
                        file_found = True
                        break
                    except FileNotFoundError:
                        continue
                
                if not file_found:
                    sftp.close()
                    ssh_client.close()
                    os.unlink(temp_path)
                    print(f"❌ Output file not found in /tmp for this run")
                    return jsonify({'error': 'Output file not found for this run'}), 404
            
            sftp.close()
            ssh_client.close()
            
            # Determine mimetype based on file extension
            mimetype = 'application/octet-stream'
            if download_filename.lower().endswith('.csv'):
                mimetype = 'text/csv'
            elif download_filename.lower().endswith('.json'):
                mimetype = 'application/json'
            elif download_filename.lower().endswith('.xml'):
                mimetype = 'application/xml'
            elif download_filename.lower().endswith('.parquet'):
                mimetype = 'application/octet-stream'
            elif download_filename.lower().endswith('.xlsx') or download_filename.lower().endswith('.xls'):
                mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            elif download_filename.lower().endswith('.tsv'):
                mimetype = 'text/tab-separated-values'
            
            # Return the file as a download
            return send_file(
                temp_path,
                mimetype=mimetype,
                as_attachment=True,
                download_name=download_filename
            )
            
        except Exception as e:
            print(f"❌ SFTP error: {str(e)}")
            return jsonify({'error': f'Failed to download file from Unix server: {str(e)}'}), 500
            
    except Exception as e:
        print(f"❌ File Download error: {str(e)}")
        return jsonify({'error': str(e)}), 500


# ============================================================
# SCHEMA GENERATION TAB - WITH FILE UPLOAD ENDPOINT
# ============================================================
@app.route('/api/static/schema/run-with-upload', methods=['POST'])
def static_schema_run_with_upload():
    """
    Execute Schema Generation operation with table list file upload.
    
    This endpoint handles multipart/form-data submissions where:
    - table_list_file: The uploaded file containing table names
    - The file is transferred to Unix /tmp/{original_filename}
    - The Unix path is then used as source input for the execution script
    
    Returns: { "run_id": "uuid", "status": "PENDING" }
    """
    try:
        # Debug logging
        print(f"[Schema Upload] Files received: {list(request.files.keys())}")
        print(f"[Schema Upload] Form fields received: {list(request.form.keys())}")
        
        # Check for uploaded file
        if 'table_list_file' not in request.files:
            return jsonify({
                'error': 'No file uploaded. Please select a table list file to upload.'
            }), 400
        
        uploaded_file = request.files['table_list_file']
        
        if uploaded_file.filename == '':
            return jsonify({
                'error': 'No file selected. Please select a table list file to upload.'
            }), 400
        
        # Read file content into memory
        file_content = uploaded_file.read()
        original_filename = uploaded_file.filename
        
        print(f"[Schema Upload] File received: {original_filename}, size: {len(file_content)} bytes")
        
        # Collect form parameters
        params = {}
        for key in request.form:
            params[key] = request.form[key]
        
        # Validate required fields
        output_file_name = params.get('output_file_name', '').strip()
        if not output_file_name:
            return jsonify({'error': 'Output File Name is required'}), 400
        
        schema_mode = params.get('schema_mode', '').strip()
        if schema_mode != 'multiple':
            return jsonify({'error': 'This endpoint is for Multiple schema mode only'}), 400
        
        # Get Unix config from params
        unix_config = get_unix_config_from_params(params)
        if 'error' in unix_config:
            return jsonify({'error': unix_config['error']}), 400
        
        # Resolve execution context
        exec_context = resolve_execution_context(params)
        if exec_context['error']:
            return jsonify({'error': exec_context['error']}), 400
        
        # Generate unique run ID
        run_id = str(uuid.uuid4())[:8]
        
        # Store file info for background thread
        # Preserve original filename (no batch_id prefix) - just sanitize special characters
        safe_filename = original_filename.replace(' ', '_').replace('/', '_').replace('\\', '_')
        unix_dest_path = f'/tmp/{safe_filename}'
        
        # Initialize run entry
        run_registry[run_id] = {
            'status': 'PENDING',
            'logs': [],
            'tab': 'schema_generation',
            'params': params,
            'thread': None,
            'ssh_client': None,
            'stop_requested': False,
            'uploaded_file_content': file_content,
            'uploaded_file_name': original_filename,
            'unix_file_path': unix_dest_path
        }
        
        # Start execution in background thread
        thread = threading.Thread(
            target=execute_schema_with_upload,
            args=(run_id, params, file_content, original_filename, unix_dest_path)
        )
        thread.daemon = True
        run_registry[run_id]['thread'] = thread
        thread.start()
        
        return jsonify({
            'run_id': run_id,
            'status': 'PENDING',
            'message': f'Schema Generation operation started with uploaded file: {original_filename}'
        })
        
    except Exception as e:
        print(f"[Schema Upload] Error: {str(e)}")
        return jsonify({'error': str(e)}), 500


def execute_schema_with_upload(run_id, params, file_content, original_filename, unix_dest_path):
    """
    Execute Schema Generation operation with table list file upload.
    
    Steps:
    1. Connect to Unix server via SSH/SFTP
    2. Upload the file to /tmp/{original_filename} (preserving original filename)
    3. Update params with the Unix file path
    4. Execute the Schema Generation script
    """
    try:
        run_registry[run_id]['status'] = 'RUNNING'
        add_log(run_id, f'🚀 Starting Schema Generation with table list upload...')
        add_log(run_id, f'📤 Uploaded file: {original_filename} ({len(file_content)} bytes)')
        
        # Resolve execution context
        exec_context = resolve_execution_context(params, run_id)
        
        if exec_context['error']:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, f'❌ {exec_context["error"]}')
            return
        
        # Extract resolved context
        working_dir = exec_context['working_dir']
        batch_id = exec_context['batch_id']
        env = exec_context['env']
        unix_config = exec_context['unix_config']
        
        # Store in run registry
        set_execution_context(run_id, working_dir, batch_id, env)
        
        add_log(run_id, f'📂 Working directory: {working_dir}')
        add_log(run_id, f'🌍 Environment: {env}, Batch ID: {batch_id}')
        add_log(run_id, f'📡 Connecting to Unix server {unix_config["host"]}:{unix_config["port"]}...')
        
        # Connect to Unix server
        ssh_client = connect_unix_server(unix_config, run_id)
        
        if not ssh_client:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, '❌ Failed to connect to Unix server')
            return
        
        run_registry[run_id]['ssh_client'] = ssh_client
        add_log(run_id, '✅ Connected to Unix server')
        
        # Step 1: Upload file to Unix /tmp via SFTP
        add_log(run_id, f'📤 Uploading table list file to Unix: {unix_dest_path}')
        
        try:
            sftp = ssh_client.open_sftp()
            
            # Write file content to Unix
            with sftp.file(unix_dest_path, 'wb') as remote_file:
                remote_file.write(file_content)
            
            sftp.close()
            add_log(run_id, f'✅ File uploaded successfully to {unix_dest_path}')
            
        except Exception as e:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, f'❌ Failed to upload file to Unix: {str(e)}')
            ssh_client.close()
            return
        
        # Step 2: Update params with Unix file path
        params['schemaSourceTableListPath'] = unix_dest_path
        
        # Check if stop requested
        if run_registry[run_id]['stop_requested']:
            ssh_client.close()
            return
        
        # Step 3: Execute the Schema Generation script
        script_path = TAB_SCRIPT_MAP['schema_generation']
        add_log(run_id, f'📋 Using script: {script_path}')
        
        # Ensure active_tab is in params
        params['active_tab'] = 'schema_generation'
        add_log(run_id, f'🏷️ Active tab passed to Unix: schema_generation')
        
        # Build script arguments
        script_args = build_script_arguments(params)
        
        # Build full command
        add_log(run_id, f'▶️ Executing script: {script_path}')
        command = f'cd {shlex.quote(working_dir)} && {script_path} {script_args}'
        
        # Execute command and stream output
        success = execute_command_streaming(ssh_client, command, run_id)
        
        # Close SSH connection
        ssh_client.close()
        run_registry[run_id]['ssh_client'] = None
        add_log(run_id, '🔌 Unix server connection closed')
        
        # Set final status
        if run_registry[run_id]['stop_requested']:
            run_registry[run_id]['status'] = 'STOPPED'
        elif success:
            run_registry[run_id]['status'] = 'SUCCESS'
            add_log(run_id, f'✅ Schema Generation completed successfully!')
            
            # Set up download info
            output_file_name = params.get('output_file_name', 'schema_output.sql')
            run_registry[run_id]['download_available'] = True
            run_registry[run_id]['download_filename'] = output_file_name
            run_registry[run_id]['unix_output_path'] = f'/tmp/{output_file_name}'
            add_log(run_id, f'⬇️ Output file ready for download: {output_file_name}')
        else:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, f'❌ Schema Generation failed')
            
    except Exception as e:
        run_registry[run_id]['status'] = 'FAILED'
        add_log(run_id, f'❌ Error: {str(e)}')
        print(f"Error in execute_schema_with_upload: {str(e)}")


@app.route('/api/static/schema/download/<run_id>')
def static_schema_download(run_id):
    """
    Download the generated schema file from Unix server.
    
    This endpoint serves the output file generated by the Schema Generation operation.
    The downloaded file name is the output_file_name specified by the user.
    """
    try:
        if run_id not in run_registry:
            return jsonify({'error': 'Run ID not found'}), 404
        
        run_entry = run_registry[run_id]
        
        # Validate this is a completed schema_generation run
        if run_entry['tab'] != 'schema_generation':
            return jsonify({'error': 'Download is only available for schema generation runs'}), 400
        
        if run_entry['status'] not in ['SUCCESS', 'PARTIAL_SUCCESS']:
            return jsonify({'error': f'Cannot download - run status is {run_entry["status"]}'}), 400
        
        # Get download info from run_entry
        if not run_entry.get('download_available'):
            return jsonify({'error': 'No download available for this run'}), 400
        
        unix_source_path = run_entry.get('unix_output_path')
        download_filename = run_entry.get('download_filename', 'schema_output.sql')
        
        if not unix_source_path:
            return jsonify({'error': 'Output file path not found in run registry'}), 500
        
        # Get Unix config from params
        params = run_entry.get('params', {})
        unix_config = get_unix_config_from_params(params)
        
        if 'error' in unix_config:
            return jsonify({'error': unix_config['error']}), 400
        
        print(f"⬇️ Preparing Schema Download from Unix: {unix_source_path} as {download_filename}")
        
        # Connect to Unix server via SFTP
        try:
            ssh_client = paramiko.SSHClient()
            ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            
            ssh_client.connect(
                hostname=unix_config['host'],
                port=unix_config['port'],
                username=unix_config['username'],
                password=unix_config['password'],
                timeout=30
            )
            
            sftp = ssh_client.open_sftp()
            
            # Create a temporary file to store the downloaded content
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.tmp')
            temp_path = temp_file.name
            temp_file.close()
            
            try:
                # Download file from Unix server
                sftp.get(unix_source_path, temp_path)
                print(f"✅ Downloaded file from {unix_source_path} to {temp_path}")
            except FileNotFoundError:
                # Try alternative paths in /tmp
                alt_paths = [
                    '/tmp/schema_output.sql',
                    '/tmp/output.sql',
                    f'/tmp/{download_filename}'
                ]
                
                file_found = False
                for alt_path in alt_paths:
                    try:
                        sftp.get(alt_path, temp_path)
                        print(f"✅ Found and downloaded file from alternative path: {alt_path}")
                        file_found = True
                        break
                    except FileNotFoundError:
                        continue
                
                if not file_found:
                    sftp.close()
                    ssh_client.close()
                    os.unlink(temp_path)
                    print(f"❌ Output file not found in /tmp for this run")
                    return jsonify({'error': 'Output file not found for this run'}), 404
            
            sftp.close()
            ssh_client.close()
            
            # Determine mimetype based on file extension
            mimetype = 'application/octet-stream'
            if download_filename.lower().endswith('.sql'):
                mimetype = 'text/plain'
            elif download_filename.lower().endswith('.csv'):
                mimetype = 'text/csv'
            elif download_filename.lower().endswith('.json'):
                mimetype = 'application/json'
            elif download_filename.lower().endswith('.txt'):
                mimetype = 'text/plain'
            
            # Return the file as a download
            return send_file(
                temp_path,
                mimetype=mimetype,
                as_attachment=True,
                download_name=download_filename
            )
            
        except Exception as e:
            print(f"❌ SFTP error: {str(e)}")
            return jsonify({'error': f'Failed to download file from Unix server: {str(e)}'}), 500
            
    except Exception as e:
        print(f"❌ Schema Download error: {str(e)}")
        return jsonify({'error': str(e)}), 500


# ============================================================
# MULTI-TABLE EXCEL COMPARISON ENDPOINT
# ============================================================
@app.route('/api/static/comparison/run-multi', methods=['POST'])
def static_run_multi_comparison():
    """
    Execute multi-table database comparison from Excel file.
    
    Content-Type: multipart/form-data
    Form fields:
      - tab = "data_comparison"
      - comparison_type = "db_to_db"
      - mode = "multiple_tables"
      - source_excel_file = uploaded .xlsx file with source table configs
      - target_excel_file = uploaded .xlsx file with target table configs
      - Other common fields (source/target DB details, environment, etc.)
    
    Returns: { "run_id": "uuid", "status": "PENDING", "total_rows": N }
    """
    try:
        # Check if openpyxl is available
        if not OPENPYXL_AVAILABLE:
            return jsonify({
                'error': 'Excel parsing not available. Please install openpyxl: pip install openpyxl'
            }), 500
        
        # Debug: Log what files were received
        print(f"[DEBUG] Files received: {list(request.files.keys())}")
        print(f"[DEBUG] Form fields received: {list(request.form.keys())}")
        
        # Validate required fields
        tab = request.form.get('tab', 'data_comparison')
        comparison_type = request.form.get('comparison_type', 'db_to_db')
        mode = request.form.get('mode', 'multiple_tables')
        
        print(f"[DEBUG] tab={tab}, comparison_type={comparison_type}, mode={mode}")
        
        if mode != 'multiple_tables':
            return jsonify({
                'error': 'This endpoint is for multiple_tables mode only'
            }), 400
        
        # Check for BOTH Excel files
        if 'source_excel_file' not in request.files:
            print(f"[DEBUG] Missing source_excel_file. Available files: {list(request.files.keys())}")
            return jsonify({
                'error': f'Both Source and Target Excel files are required for multiple_tables mode. Missing: source_excel_file. Received files: {list(request.files.keys())}'
            }), 400
        
        if 'target_excel_file' not in request.files:
            print(f"[DEBUG] Missing target_excel_file. Available files: {list(request.files.keys())}")
            return jsonify({
                'error': f'Both Source and Target Excel files are required for multiple_tables mode. Missing: target_excel_file. Received files: {list(request.files.keys())}'
            }), 400
        
        source_excel_file = request.files['source_excel_file']
        target_excel_file = request.files['target_excel_file']
        
        # Log file info
        print(f"[DEBUG] Source file: {source_excel_file.filename}, size={source_excel_file.content_length}")
        print(f"[DEBUG] Target file: {target_excel_file.filename}, size={target_excel_file.content_length}")
        
        if source_excel_file.filename == '':
            return jsonify({
                'error': 'No source file selected'
            }), 400
        
        if target_excel_file.filename == '':
            return jsonify({
                'error': 'No target file selected'
            }), 400
        
        if not source_excel_file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({
                'error': 'Invalid source file type. Please upload an Excel file (.xlsx or .xls)'
            }), 400
        
        if not target_excel_file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({
                'error': 'Invalid target file type. Please upload an Excel file (.xlsx or .xls)'
            }), 400
        
        # Read both Excel file contents
        try:
            source_content = source_excel_file.read()
            target_content = target_excel_file.read()
            
            print(f"[DEBUG] Source content size: {len(source_content)} bytes")
            print(f"[DEBUG] Target content size: {len(target_content)} bytes")
            
            source_rows = parse_excel_for_multi_comparison(source_content, 'source')
            target_rows = parse_excel_for_multi_comparison(target_content, 'target')
            
            print(f"[DEBUG] Parsed source_rows: {len(source_rows)}")
            print(f"[DEBUG] Parsed target_rows: {len(target_rows)}")
            
            # Log first row of each for debugging
            if source_rows:
                print(f"[DEBUG] First source row: {source_rows[0]}")
            if target_rows:
                print(f"[DEBUG] First target row: {target_rows[0]}")
                
        except Exception as e:
            print(f"[DEBUG] Excel parsing error: {str(e)}")
            return jsonify({
                'error': f'Failed to parse Excel file: {str(e)}'
            }), 400
        
        if not source_rows:
            return jsonify({
                'error': 'Source Excel file is empty or has no valid rows'
            }), 400
        
        if not target_rows:
            return jsonify({
                'error': 'Target Excel file is empty or has no valid rows'
            }), 400
        
        # Collect common form parameters (excluding file)
        common_params = {}
        for key in request.form:
            if key not in ['tab', 'comparison_type', 'mode', 'source_excel_file', 'target_excel_file']:
                common_params[key] = request.form[key]
        
        # Generate unique run ID
        run_id = str(uuid.uuid4())[:8]
        
        # Determine rows to process (min of both)
        total_rows = min(len(source_rows), len(target_rows))
        
        # Initialize run entry with multi-table specific fields
        run_registry[run_id] = {
            'status': 'PENDING',
            'logs': [],
            'tab': tab,
            'params': common_params,
            'thread': None,
            'ssh_client': None,
            'stop_requested': False,
            'rows_status': [],
            'total_rows': total_rows,
            'started_at': datetime.now().isoformat(),
            'finished_at': None,
            'source_rows': source_rows,
            'target_rows': target_rows
        }
        
        # Start execution in background thread
        thread = threading.Thread(
            target=execute_multi_table_comparison_v2,
            args=(run_id, common_params, source_rows, target_rows)
        )
        thread.daemon = True
        run_registry[run_id]['thread'] = thread
        thread.start()
        
        return jsonify({
            'run_id': run_id,
            'status': 'PENDING',
            'total_rows': total_rows,
            'source_rows': len(source_rows),
            'target_rows': len(target_rows),
            'message': f'Multi-table comparison started with {total_rows} rows (source: {len(source_rows)}, target: {len(target_rows)})'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/static/run/<run_id>/summary')
def get_run_summary(run_id):
    """
    Get execution summary for a completed run.
    Returns row-wise results for multi-table comparisons.
    """
    if run_id not in run_registry:
        return jsonify({'error': 'Run ID not found'}), 404
    
    run_entry = run_registry[run_id]
    
    # Calculate success/failure counts
    rows_status = run_entry.get('rows_status', [])
    success_count = sum(1 for r in rows_status if r.get('status') == 'SUCCESS')
    failed_count = sum(1 for r in rows_status if r.get('status') == 'FAILED')
    
    return jsonify({
        'run_id': run_id,
        'status': run_entry['status'],
        'tab': run_entry['tab'],
        'total_rows': run_entry.get('total_rows', 0),
        'success_count': success_count,
        'failed_count': failed_count,
        'rows_status': rows_status,
        'started_at': run_entry.get('started_at'),
        'finished_at': run_entry.get('finished_at')
    })


# ============================================================
# EXCEL PARSING FOR MULTI-TABLE COMPARISON
# ============================================================
def normalize_column_name(name):
    """
    Normalize column name: trim spaces, lowercase, replace spaces with underscores.
    """
    if not name:
        return ''
    return str(name).strip().lower().replace(' ', '_').replace('-', '_')


def parse_excel_for_multi_comparison(excel_content, file_type='source'):
    """
    Parse Excel file for multi-table comparison.
    
    file_type: 'source' or 'target'
    
    For SOURCE file, expects columns:
    - source_table_name (or aliases: table_name, src_table_name, src_table)
    - source_sql (or aliases: source_sql_query, sql_query, sql, src_sql)
    - source_key_columns (or aliases: key_columns, source_keys, src_keys)
    
    For TARGET file, expects columns:
    - target_table_name (or aliases: table_name, tgt_table_name, tgt_table)
    - target_sql (or aliases: target_sql_query, sql_query, sql, tgt_sql)
    - target_key_columns (or aliases: key_columns, target_keys, tgt_keys)
    
    Returns: List of row dictionaries with standardized keys
    """
    workbook = openpyxl.load_workbook(BytesIO(excel_content), read_only=True, data_only=True)
    sheet = workbook.active
    
    # Get headers from first row
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    
    if not header_row:
        raise ValueError("Excel file has no header row")
    
    # Normalize headers
    headers = [normalize_column_name(h) for h in header_row]
    
    print(f"[DEBUG] {file_type.upper()} Excel headers (normalized): {headers}")
    
    # Define column aliases based on file type
    if file_type == 'source':
        column_aliases = {
            'table_name': ['source_table_name', 'src_table_name', 'src_table', 'table_name', 'tablename', 'source_table'],
            'sql_query': ['source_sql', 'source_sql_query', 'src_sql', 'src_sql_query', 'sql_query', 'sql', 'query'],
            'key_columns': ['source_key_columns', 'source_keys', 'src_keys', 'src_key_columns', 'key_columns', 'keys', 'primary_key']
        }
    else:  # target
        column_aliases = {
            'table_name': ['target_table_name', 'tgt_table_name', 'tgt_table', 'table_name', 'tablename', 'target_table'],
            'sql_query': ['target_sql', 'target_sql_query', 'tgt_sql', 'tgt_sql_query', 'sql_query', 'sql', 'query'],
            'key_columns': ['target_key_columns', 'target_keys', 'tgt_keys', 'tgt_key_columns', 'key_columns', 'keys', 'primary_key']
        }
    
    # Build header index map
    header_map = {}
    missing_columns = []
    
    for canonical_name, aliases in column_aliases.items():
        found = False
        for alias in aliases:
            if alias in headers:
                header_map[canonical_name] = headers.index(alias)
                found = True
                break
        if not found:
            missing_columns.append(f"{canonical_name} (or aliases: {', '.join(aliases[:3])})")
    
    print(f"[DEBUG] {file_type.upper()} Excel header_map: {header_map}")
    if missing_columns:
        print(f"[DEBUG] {file_type.upper()} Excel missing columns: {missing_columns}")
    
    # Warn about missing columns but don't fail if table_name is present
    if 'table_name' not in header_map:
        raise ValueError(f"Missing required columns in {file_type} Excel file: {', '.join(missing_columns)}. Detected headers: {headers}")
    
    # Parse data rows
    rows = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        
        row_data = {
            'row_index': row_num - 1,  # 1-based index for display
            'table_name': '',
            'sql_query': '',
            'key_columns': ''
        }
        
        # Extract values using header map
        for canonical_name, col_index in header_map.items():
            if col_index < len(row) and row[col_index] is not None:
                row_data[canonical_name] = str(row[col_index]).strip()
        
        # Skip rows where table_name is empty
        if not row_data['table_name']:
            continue
        
        rows.append(row_data)
    
    workbook.close()
    print(f"[DEBUG] {file_type.upper()} Excel: parsed {len(rows)} valid data rows")
    return rows


def parse_excel_for_comparison(excel_content):
    """
    Parse Excel file and extract rows for comparison.
    
    Expected columns (case-insensitive, supports alternate names):
    - source_sql_query / source_query / source_sql
    - source_table_name / source_table / table_name
    - source_key_columns / source_keys / key_columns
    - target_sql_query / target_query / target_sql
    - target_table_name / target_table
    - target_key_columns / target_keys
    
    Returns: List of row dictionaries
    """
    workbook = openpyxl.load_workbook(BytesIO(excel_content), read_only=True, data_only=True)
    sheet = workbook.active
    
    # Get headers from first row
    headers = []
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    
    if not header_row:
        raise ValueError("Excel file has no header row")
    
    # Normalize headers (lowercase, strip whitespace)
    headers = [str(h).lower().strip() if h else '' for h in header_row]
    
    # Column name mapping (alternate names -> canonical name)
    column_aliases = {
        'source_sql_query': ['source_sql_query', 'source_query', 'source_sql', 'src_query', 'src_sql'],
        'source_table_name': ['source_table_name', 'source_table', 'src_table', 'src_table_name'],
        'source_key_columns': ['source_key_columns', 'source_keys', 'src_keys', 'src_key_columns', 'key_columns'],
        'target_sql_query': ['target_sql_query', 'target_query', 'target_sql', 'tgt_query', 'tgt_sql'],
        'target_table_name': ['target_table_name', 'target_table', 'tgt_table', 'tgt_table_name'],
        'target_key_columns': ['target_key_columns', 'target_keys', 'tgt_keys', 'tgt_key_columns']
    }
    
    # Build header index map
    header_map = {}
    for canonical_name, aliases in column_aliases.items():
        for alias in aliases:
            if alias in headers:
                header_map[canonical_name] = headers.index(alias)
                break
    
    # Parse data rows
    rows = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        
        row_data = {
            'row_index': row_num - 1,  # 1-based index for display
            'source_sql_query': '',
            'source_table_name': '',
            'source_key_columns': '',
            'target_sql_query': '',
            'target_table_name': '',
            'target_key_columns': ''
        }
        
        # Extract values using header map
        for canonical_name, col_index in header_map.items():
            if col_index < len(row) and row[col_index] is not None:
                row_data[canonical_name] = str(row[col_index]).strip()
        
        # Also capture any additional columns not in the mapping
        for i, header in enumerate(headers):
            if header and header not in ['', None] and i < len(row):
                normalized_header = header.replace(' ', '_').replace('-', '_')
                if normalized_header not in row_data and row[i] is not None:
                    row_data[normalized_header] = str(row[i]).strip()
        
        rows.append(row_data)
    
    workbook.close()
    return rows


# ============================================================
# MULTI-TABLE COMPARISON EXECUTION
# ============================================================
def execute_multi_table_comparison_v2(run_id, common_params, source_rows, target_rows):
    """
    Execute multi-table comparison by iterating through source and target Excel rows.
    Each row triggers a Unix script execution via Paramiko.
    """
    try:
        run_registry[run_id]['status'] = 'RUNNING'
        num_source_rows = len(source_rows)
        num_target_rows = len(target_rows)
        total_rows = min(num_source_rows, num_target_rows)
        
        add_log(run_id, f'🚀 Starting Multi-Table Comparison...')
        add_log(run_id, f'📊 Parsed source rows: {num_source_rows}, Parsed target rows: {num_target_rows}, Processing rows: {total_rows}')
        
        if num_source_rows != num_target_rows:
            add_log(run_id, f'⚠️ WARNING: Row count mismatch! Source has {num_source_rows} rows, Target has {num_target_rows} rows. Will process {total_rows} rows.')
        
        add_log(run_id, '=' * 60)
        
        # Log preview of first 2 rows for debugging
        for i in range(min(2, total_rows)):
            src = source_rows[i]
            tgt = target_rows[i]
            add_log(run_id, f'📋 Preview Row {i+1}:')
            add_log(run_id, f'   Source: table={src.get("table_name", "N/A")}, keys={src.get("key_columns", "N/A")}, sql={src.get("sql_query", "N/A")[:50]}...')
            add_log(run_id, f'   Target: table={tgt.get("table_name", "N/A")}, keys={tgt.get("key_columns", "N/A")}, sql={tgt.get("sql_query", "N/A")[:50]}...')
        
        add_log(run_id, '=' * 60)
        
        # Resolve execution context ONCE using central function
        exec_context = resolve_execution_context(common_params, run_id)
        
        if exec_context['error']:
            run_registry[run_id]['status'] = 'FAILED'
            run_registry[run_id]['finished_at'] = datetime.now().isoformat()
            add_log(run_id, f'❌ {exec_context["error"]}')
            return
        
        # Extract resolved context
        working_dir = exec_context['working_dir']
        batch_id = exec_context['batch_id']
        env = exec_context['env']
        unix_config = exec_context['unix_config']
        
        # Store in run registry for global access
        set_execution_context(run_id, working_dir, batch_id, env)
        
        # Log working directory ONCE per run
        add_log(run_id, f'📂 Global working directory set to: {working_dir}')
        add_log(run_id, f'🌍 Environment: {env}, Batch ID: {batch_id}')
        add_log(run_id, f'🔧 Unix host: {unix_config["host"]}')
        add_log(run_id, f'🏷️ Active tab passed to Unix: data_comparison')
        
        # Extract and log time_zone (default to N/A if not provided)
        time_zone = extract_time_zone(common_params)
        common_params['time_zone'] = time_zone  # Ensure it's in params for script args
        add_log(run_id, f'🕐 Time Zone: {time_zone}')
        
        add_log(run_id, f'📡 Connecting to Unix server {unix_config["host"]}:{unix_config["port"]}...')
        
        # Connect to Unix server
        ssh_client = connect_unix_server(unix_config, run_id)
        
        if not ssh_client:
            run_registry[run_id]['status'] = 'FAILED'
            run_registry[run_id]['finished_at'] = datetime.now().isoformat()
            add_log(run_id, '❌ Failed to connect to Unix server')
            return
        
        run_registry[run_id]['ssh_client'] = ssh_client
        add_log(run_id, '✅ Successfully connected to Unix server')
        add_log(run_id, '=' * 60)
        
        success_count = 0
        failed_count = 0
        
        # Process each row
        for i in range(total_rows):
            row_num = i + 1
            source_row = source_rows[i]
            target_row = target_rows[i]
            
            # Check if stop requested
            if run_registry[run_id]['stop_requested']:
                add_log(run_id, f'⛔ Execution stopped by user at row {row_num}/{total_rows}')
                break
            
            source_table = source_row.get('table_name', 'unknown')
            source_sql = source_row.get('sql_query', '')
            source_keys = source_row.get('key_columns', '')
            
            target_table = target_row.get('table_name', 'unknown')
            target_sql = target_row.get('sql_query', '')
            target_keys = target_row.get('key_columns', '')
            
            add_log(run_id, '')
            add_log(run_id, f'[Row {row_num}/{total_rows}] START')
            add_log(run_id, f'[Row {row_num}/{total_rows}] Source table={source_table}, keys={source_keys}, sql={source_sql[:80] if source_sql else "N/A"}...')
            add_log(run_id, f'[Row {row_num}/{total_rows}] Target table={target_table}, keys={target_keys}, sql={target_sql[:80] if target_sql else "N/A"}...')
            
            row_status = {
                'row_index': row_num,
                'source_table': source_table,
                'target_table': target_table,
                'source_keys': source_keys,
                'target_keys': target_keys,
                'status': 'RUNNING',
                'error': None
            }
            
            try:
                # Build combined parameters: common UI fields + Excel row fields
                combined_params = {**common_params}
                
                # Add row-specific fields from both source and target
                combined_params['source_sql_query'] = source_sql
                combined_params['source_table_name'] = source_table
                combined_params['source_key_columns'] = source_keys
                combined_params['target_sql_query'] = target_sql
                combined_params['target_table_name'] = target_table
                combined_params['target_key_columns'] = target_keys
                combined_params['row_index'] = row_num
                combined_params['total_rows'] = total_rows
                
                # Use MULTI_TABLE_SCRIPT (global constant)
                script_path = MULTI_TABLE_SCRIPT
                
                # Ensure active_tab is in params
                combined_params['active_tab'] = 'data_comparison'
                
                script_args = build_script_arguments(combined_params)
                
                # Build full command using global working_dir (no redundant log)
                add_log(run_id, f'  ▶️ Executing script: {script_path}')
                command = f'cd {shlex.quote(working_dir)} && {script_path} {script_args}'
                
                # Execute command
                success = execute_command_streaming(ssh_client, command, run_id)
                
                if success:
                    row_status['status'] = 'SUCCESS'
                    success_count += 1
                    add_log(run_id, f'[Row {row_num}/{total_rows}] END status=SUCCESS ✅')
                else:
                    row_status['status'] = 'FAILED'
                    row_status['error'] = 'Script execution failed'
                    failed_count += 1
                    add_log(run_id, f'[Row {row_num}/{total_rows}] END status=FAILED ❌')
                
            except Exception as e:
                row_status['status'] = 'FAILED'
                row_status['error'] = str(e)
                failed_count += 1
                add_log(run_id, f'  ❌ Error: {str(e)}')
                add_log(run_id, f'[Row {row_num}/{total_rows}] END status=FAILED ❌')
            
            # Record row status
            run_registry[run_id]['rows_status'].append(row_status)
        
        # Close SSH connection
        ssh_client.close()
        run_registry[run_id]['ssh_client'] = None
        
        add_log(run_id, '')
        add_log(run_id, '=' * 60)
        add_log(run_id, '🔌 Unix server connection closed')
        
        # Determine final status
        run_registry[run_id]['finished_at'] = datetime.now().isoformat()
        
        if run_registry[run_id]['stop_requested']:
            run_registry[run_id]['status'] = 'STOPPED'
            add_log(run_id, f'⛔ Execution stopped. Processed: {success_count + failed_count}/{total_rows}')
        elif failed_count == 0:
            run_registry[run_id]['status'] = 'SUCCESS'
            add_log(run_id, f'✅ All {total_rows} comparisons completed successfully!')
        elif success_count == 0:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, f'❌ All {total_rows} comparisons failed')
        else:
            run_registry[run_id]['status'] = 'PARTIAL_SUCCESS'
            add_log(run_id, f'⚠️ Completed with partial success: {success_count} succeeded, {failed_count} failed')
        
        add_log(run_id, '=' * 60)
        add_log(run_id, f'📊 Summary: {success_count} success, {failed_count} failed, {total_rows} total')
        
    except Exception as e:
        run_registry[run_id]['status'] = 'FAILED'
        run_registry[run_id]['finished_at'] = datetime.now().isoformat()
        add_log(run_id, f'❌ Fatal error: {str(e)}')
        print(f"Error in execute_multi_table_comparison: {str(e)}")


def execute_multi_table_comparison(run_id, common_params, excel_rows):
    """
    Legacy: Execute multi-table comparison by iterating through Excel rows.
    Kept for backwards compatibility with single-file uploads.
    """
    try:
        run_registry[run_id]['status'] = 'RUNNING'
        total_rows = len(excel_rows)
        
        add_log(run_id, f'🚀 Starting Multi-Table Comparison ({total_rows} tables)...')
        add_log(run_id, '=' * 60)
        
        # Get Unix config from params (UI-provided values override env vars)
        unix_config = get_unix_config_from_params(common_params)
        
        # Check for missing host error
        if 'error' in unix_config:
            run_registry[run_id]['status'] = 'FAILED'
            run_registry[run_id]['finished_at'] = datetime.now().isoformat()
            add_log(run_id, f'❌ {unix_config["error"]}')
            return
        
        add_log(run_id, f'📊 Parsed rows: {total_rows}')
        add_log(run_id, f'❌ Legacy single-file mode is deprecated. Please use two-file upload.')
        run_registry[run_id]['status'] = 'FAILED'
        run_registry[run_id]['finished_at'] = datetime.now().isoformat()
        
    except Exception as e:
        run_registry[run_id]['status'] = 'FAILED'
        run_registry[run_id]['finished_at'] = datetime.now().isoformat()
        add_log(run_id, f'❌ Fatal error: {str(e)}')


# ============================================================
# EXECUTION LOGIC
# ============================================================
def execute_static_operation(run_id, tab, params):
    """
    Execute operation in background thread with real-time logging.
    
    Uses centralized working_dir from resolve_execution_context().
    """
    try:
        run_registry[run_id]['status'] = 'RUNNING'
        add_log(run_id, f'🚀 Starting {tab.replace("_", " ").title()} operation...')
        
        # Determine comparison mode for data_comparison tab
        comparison_mode = normalize_comparison_mode(params)
        if tab == 'data_comparison':
            add_log(run_id, f'📊 Mode: {comparison_mode.replace("_", " ").title()}')
            
            # For multiple_tables mode via /api/static/run endpoint without Excel files,
            # we cannot iterate - Excel files are required
            if comparison_mode == 'multiple_tables':
                run_registry[run_id]['status'] = 'FAILED'
                add_log(run_id, '❌ Multiple Tables mode requires Excel file uploads.')
                add_log(run_id, '❌ Please use the /api/static/comparison/run-multi endpoint with source and target Excel files.')
                return
        
        # Data Load tab specific validation and field extraction
        if tab == 'data_load':
            # Validate Data Load specific params
            is_valid, error_msg = validate_data_load_params(params)
            if not is_valid:
                run_registry[run_id]['status'] = 'FAILED'
                add_log(run_id, f'❌ Validation Error: {error_msg}')
                return
            
            # Extract and add Data Load fields to params
            load_fields = extract_data_load_fields(params)
            params['target_load_type'] = load_fields['target_load_type']
            add_log(run_id, f'⚡ Load Type: {load_fields["target_load_type"]}')
            
            # Add Hive-specific fields if applicable
            if load_fields['target_db_type'].lower() == 'hive':
                params['hadoop_path'] = load_fields['hadoop_path']
                params['partition_id'] = load_fields['partition_id']
                add_log(run_id, f'🐝 Target DB: Hive')
                add_log(run_id, f'📂 Hadoop Path: {load_fields["hadoop_path"]}')
                add_log(run_id, f'🔖 Partition ID: {load_fields["partition_id"]}')
        
        # File Load tab specific validation and field extraction (File-to-Database)
        if tab == 'file_load':
            # Validate File Load specific params
            is_valid, error_msg = validate_file_load_params(params)
            if not is_valid:
                run_registry[run_id]['status'] = 'FAILED'
                add_log(run_id, f'❌ Validation Error: {error_msg}')
                return
            
            # Extract and add File Load source fields to params
            source_fields = extract_file_load_source_fields(params)
            params['source_file_type'] = source_fields['source_file_type']
            params['source_location_type'] = source_fields['source_location_type']
            params['source_unix_path'] = source_fields['source_unix_path']
            params['source_hadoop_path'] = source_fields['source_hadoop_path']
            params['source_delimiter'] = source_fields['source_delimiter']
            params['source_file_sql_query'] = source_fields['source_file_sql_query']
            params['source_key_columns'] = source_fields['source_key_columns']
            
            add_log(run_id, f'📄 Source File Type: {source_fields["source_file_type"]}')
            add_log(run_id, f'📍 Source Location: {source_fields["source_location_type"]}')
            if source_fields['source_unix_path']:
                add_log(run_id, f'📂 Source Unix Path: {source_fields["source_unix_path"]}')
            if source_fields['source_hadoop_path']:
                add_log(run_id, f'📂 Source Hadoop Path: {source_fields["source_hadoop_path"]}')
            
            # Extract and add File Load target fields (same as Data Load)
            target_fields = extract_file_load_target_fields(params)
            params['target_load_type'] = target_fields['target_load_type']
            add_log(run_id, f'⚡ Target Load Type: {target_fields["target_load_type"]}')
            
            # Add Hive-specific fields if applicable
            if target_fields['target_db_type'].lower() == 'hive':
                params['hadoop_path'] = target_fields['hadoop_path']
                params['partition_id'] = target_fields['partition_id']
                add_log(run_id, f'🐝 Target DB: Hive')
                add_log(run_id, f'📂 Hadoop Path: {target_fields["hadoop_path"]}')
                add_log(run_id, f'🔖 Partition ID: {target_fields["partition_id"]}')
        
        # File Download tab specific validation and field extraction
        if tab == 'file_download':
            # Validate File Download specific params
            is_valid, error_msg = validate_file_download_params(params)
            if not is_valid:
                run_registry[run_id]['status'] = 'FAILED'
                add_log(run_id, f'❌ Validation Error: {error_msg}')
                return
            
            # Extract and log File Download fields
            download_fields = extract_file_download_fields(params)
            params['source_type'] = download_fields['source_type']
            params['target_file_type'] = download_fields['target_file_type']
            params['target_file_name'] = download_fields['target_file_name']
            
            if download_fields['target_delimiter']:
                params['target_delimiter'] = download_fields['target_delimiter']
            if download_fields['number_of_rows']:
                params['number_of_rows'] = download_fields['number_of_rows']
            
            add_log(run_id, f'📥 Source Type: {download_fields["source_type"]}')
            add_log(run_id, f'📄 Target File Type: {download_fields["target_file_type"]}')
            add_log(run_id, f'📁 Target File Name: {download_fields["target_file_name"]}')
            if download_fields['number_of_rows']:
                add_log(run_id, f'🔢 Number of Rows: {download_fields["number_of_rows"]}')
        
        # Resolve execution context ONCE using central function
        exec_context = resolve_execution_context(params, run_id)
        
        if exec_context['error']:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, f'❌ {exec_context["error"]}')
            return
        
        # Extract resolved context
        working_dir = exec_context['working_dir']
        batch_id = exec_context['batch_id']
        env = exec_context['env']
        unix_config = exec_context['unix_config']
        
        # Store in run registry for global access
        set_execution_context(run_id, working_dir, batch_id, env)
        
        # Log working directory ONCE per run
        add_log(run_id, f'📂 Global working directory set to: {working_dir}')
        add_log(run_id, f'🌍 Environment: {env}, Batch ID: {batch_id}')
        add_log(run_id, f'🔧 Unix host: {unix_config["host"]}')
        
        # Extract and log time_zone (default to N/A if not provided)
        time_zone = extract_time_zone(params)
        params['time_zone'] = time_zone  # Ensure it's in params for script args
        add_log(run_id, f'🕐 Time Zone: {time_zone}')
        
        add_log(run_id, f'📡 Connecting to Unix server {unix_config["host"]}:{unix_config["port"]}...')
        
        # Connect to Unix server
        ssh_client = connect_unix_server(unix_config, run_id)
        
        if not ssh_client:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, '❌ Failed to connect to Unix server')
            return
        
        run_registry[run_id]['ssh_client'] = ssh_client
        add_log(run_id, '✅ Successfully connected to Unix server')
        
        # Check if stop requested
        if run_registry[run_id]['stop_requested']:
            ssh_client.close()
            return
        
        # Determine script path based on tab
        # For data_comparison single_table mode: use SINGLE_TABLE_SCRIPT
        # For other tabs: use TAB_SCRIPT_MAP
        if tab == 'data_comparison':
            # Single table mode only (multiple_tables is handled above with early return)
            script_path = SINGLE_TABLE_SCRIPT
            add_log(run_id, f'📋 Using single-table script: {script_path}')
        else:
            # Other tabs - use TAB_SCRIPT_MAP
            script_path = TAB_SCRIPT_MAP[tab]
            add_log(run_id, f'📋 Using script: {script_path}')
        
        # Ensure active_tab is in params
        params['active_tab'] = tab
        add_log(run_id, f'🏷️ Active tab passed to Unix: {tab}')
        
        # Build script arguments
        script_args = build_script_arguments(params)
        
        # Build full command using global working_dir (no redundant log)
        add_log(run_id, f'▶️ Executing script: {script_path}')
        command = f'cd {shlex.quote(working_dir)} && {script_path} {script_args}'
        
        # Execute command and stream output
        success = execute_command_streaming(ssh_client, command, run_id)
        
        # Close SSH connection
        ssh_client.close()
        run_registry[run_id]['ssh_client'] = None
        add_log(run_id, '🔌 Unix server connection closed')
        
        # Set final status
        if run_registry[run_id]['stop_requested']:
            run_registry[run_id]['status'] = 'STOPPED'
        elif success:
            run_registry[run_id]['status'] = 'SUCCESS'
            add_log(run_id, f'✅ {tab.replace("_", " ").title()} completed successfully!')
            
            # For single_table data_comparison, set up download info
            if tab == 'data_comparison' and comparison_mode == 'single_table':
                download_info = determine_download_filename(params)
                run_registry[run_id]['download_available'] = True
                run_registry[run_id]['download_filename'] = download_info['filename']
                run_registry[run_id]['unix_output_path'] = download_info['unix_path']
                add_log(run_id, f'⬇️ Output file ready for download: {download_info["filename"]}')
            
            # For file_download tab, set up download info with user-specified filename
            if tab == 'file_download':
                download_fields = extract_file_download_fields(params)
                target_file_name = download_fields['target_file_name']
                run_registry[run_id]['download_available'] = True
                run_registry[run_id]['download_filename'] = target_file_name
                run_registry[run_id]['unix_output_path'] = '/tmp/output.csv'  # Default path, script should output here
                add_log(run_id, f'⬇️ Output file ready for download: {target_file_name}')
            
            # For schema_generation tab, set up download info with user-specified output filename
            if tab == 'schema_generation':
                output_file_name = (
                    params.get('output_file_name') or
                    params.get('schemaTargetOutputFileName') or
                    'schema_output.sql'
                ).strip()
                run_registry[run_id]['download_available'] = True
                run_registry[run_id]['download_filename'] = output_file_name
                run_registry[run_id]['unix_output_path'] = f'/tmp/{output_file_name}'
                add_log(run_id, f'⬇️ Output file ready for download: {output_file_name}')
        else:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, f'❌ {tab.replace("_", " ").title()} failed')
            
    except Exception as e:
        run_registry[run_id]['status'] = 'FAILED'
        add_log(run_id, f'❌ Error: {str(e)}')
        print(f"Error in execute_static_operation: {str(e)}")

def connect_unix_server(unix_config, run_id=None):
    """Connect to Unix server using Paramiko with environment-based config"""
    try:
        ssh_client = paramiko.SSHClient()
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        
        connect_kwargs = {
            'hostname': unix_config['host'],
            'port': unix_config['port'],
            'username': unix_config['username'],
            'timeout': 30
        }
        
        # Always use password authentication (SSH key auth removed)
        connect_kwargs['password'] = unix_config['password']
        if run_id:
            add_log(run_id, '🔐 Using password authentication')
        
        ssh_client.connect(**connect_kwargs)
        return ssh_client
        
    except Exception as e:
        if run_id:
            add_log(run_id, f'❌ SSH connection error: {str(e)}')
        print(f"SSH connection error: {str(e)}")
        return None

def execute_command_streaming(ssh_client, command, run_id):
    """Execute command and stream output to logs"""
    try:
        add_log(run_id, f'⚙️ Running command...')
        
        # Execute command
        stdin, stdout, stderr = ssh_client.exec_command(command)
        
        # Stream stdout line by line
        for line in iter(stdout.readline, ''):
            if run_registry[run_id]['stop_requested']:
                return False
            if line:
                add_log(run_id, line.strip())
        
        # Check stderr
        stderr_output = stderr.read().decode()
        if stderr_output:
            for line in stderr_output.split('\n'):
                if line.strip():
                    add_log(run_id, f'⚠️ {line.strip()}')
        
        # Get exit status
        exit_status = stdout.channel.recv_exit_status()
        
        if exit_status == 0:
            add_log(run_id, '✅ Command executed successfully')
            return True
        else:
            add_log(run_id, f'❌ Command failed with exit code {exit_status}')
            return False
            
    except Exception as e:
        add_log(run_id, f'❌ Command execution error: {str(e)}')
        return False

def escape_shell_arg(arg):
    """Safely escape shell arguments to handle special characters"""
    if arg is None:
        return "''"
    
    arg_str = str(arg)
    return shlex.quote(arg_str)

def build_script_arguments(params):
    """Build properly escaped command-line arguments from params dict"""
    arg_parts = []
    
    # First, extract file path fields if this is a file-based comparison
    file_path_fields = extract_file_path_fields(params)
    
    # Merge file path fields into params for processing
    merged_params = {**params, **file_path_fields}
    
    for key, value in merged_params.items():
        if value is not None and value != '':
            escaped_value = escape_shell_arg(value)
            # Convert camelCase to snake_case for CLI args
            cli_key = ''.join(['_' + c.lower() if c.isupper() else c for c in key]).lstrip('_')
            arg_parts.append(f'--{cli_key}={escaped_value}')
    
    return ' '.join(arg_parts)

def build_command_with_args(script_path, params):
    """Build shell command with properly escaped arguments from params dict"""
    command_parts = [script_path]
    
    for key, value in params.items():
        if value is not None and value != '':
            escaped_value = escape_shell_arg(value)
            # Convert camelCase to snake_case for CLI args
            cli_key = ''.join(['_' + c.lower() if c.isupper() else c for c in key]).lstrip('_')
            command_parts.append(f'--{cli_key}={escaped_value}')
    
    return ' '.join(command_parts)

def add_log(run_id, message):
    """Add log message to run registry"""
    if run_id in run_registry:
        timestamp = time.strftime('%H:%M:%S')
        log_entry = f'[{timestamp}] {message}'
        run_registry[run_id]['logs'].append(log_entry)
        print(log_entry)

# ============================================================
# LEGACY WEBSOCKET ENDPOINTS (kept for backwards compatibility)
# ============================================================
@app.route('/api/execute-operation', methods=['POST'])
def execute_operation():
    data = request.json
    operation_type = data.get('operation', 'Unknown')
    
    print(f"Operation requested - Type: {operation_type}")
    print(f"Operation data: {data}")
    
    # Start operation in background thread
    thread = threading.Thread(target=process_operation, args=(data, operation_type))
    thread.daemon = True
    thread.start()
    
    return jsonify({"status": "success", "message": f"{operation_type} operation started"})

def process_operation(data, operation_type):
    """Process operation in background thread with real-time logging (legacy WebSocket)"""
    try:
        # Extract Unix server connection details from form data (legacy)
        unix_config = extract_unix_config(data)
        
        # Extract source and target configurations
        source_config = extract_source_config(data, operation_type)
        target_config = extract_target_config(data, operation_type)
        
        emit_log(f"Connecting to Unix server {unix_config['host']}:{unix_config['port']}...")
        ssh_client = connect_unix_server(unix_config)
        
        if not ssh_client:
            emit_error("Failed to connect to Unix server")
            return
        
        emit_log("Successfully connected to Unix server")
        
        # Execute operation based on type
        if operation_type == 'Data Comparison':
            execute_data_comparison(ssh_client, source_config, target_config, data)
        elif operation_type == 'Data Load':
            execute_data_load(ssh_client, source_config, target_config, data)
        elif operation_type == 'Schema Generation':
            execute_schema_generation(ssh_client, source_config, target_config, data)
        elif operation_type == 'File Load':
            execute_file_load(ssh_client, source_config, target_config, data)
        elif operation_type == 'File Download':
            execute_file_download(ssh_client, source_config, target_config, data)
        else:
            emit_error(f"Unknown operation type: {operation_type}")
        
        ssh_client.close()
        emit_log("Unix server connection closed")
        
    except Exception as e:
        emit_error(f"Operation failed: {str(e)}")
        print(f"Error in process_operation: {str(e)}")

def execute_command(ssh_client, command):
    """Execute command on Unix server and stream output (legacy)"""
    try:
        emit_log(f"Executing command: {command}")
        
        stdin, stdout, stderr = ssh_client.exec_command(command)
        
        for line in stdout:
            emit_log(line.strip())
        
        stderr_output = stderr.read().decode()
        if stderr_output:
            emit_log(f"Warnings/Errors: {stderr_output}")
        
        exit_status = stdout.channel.recv_exit_status()
        
        if exit_status == 0:
            emit_log("Command executed successfully")
            return True
        else:
            emit_error(f"Command failed with exit status {exit_status}")
            return False
            
    except Exception as e:
        emit_error(f"Command execution error: {str(e)}")
        return False

def execute_data_comparison(ssh_client, source_config, target_config, data):
    """Execute data comparison operation (legacy)"""
    try:
        emit_log("Starting data comparison operation...")
        
        comparison_type = data.get('comparison-type-select', 'database-to-database')
        table_mode = data.get('tableMode', 'single')
        key_columns = data.get('keyColumns', '')
        
        emit_log(f"Comparison Type: {comparison_type}")
        emit_log(f"Table Mode: {table_mode}")
        
        args = {
            'comparison_type': comparison_type,
            'table_mode': table_mode,
            'key_columns': key_columns,
            'source_type': source_config.get('type'),
            'source_host': source_config.get('host'),
            'source_port': source_config.get('port'),
            'source_database': source_config.get('database'),
            'source_username': source_config.get('username'),
            'source_password': source_config.get('password'),
            'source_table': source_config.get('table'),
            'source_sql': source_config.get('sql'),
            'source_file_type': source_config.get('file_type'),
            'source_delimiter': source_config.get('delimiter'),
            'source_file_path': source_config.get('file_path'),
            'target_type': target_config.get('type'),
            'target_host': target_config.get('host'),
            'target_port': target_config.get('port'),
            'target_database': target_config.get('database'),
            'target_username': target_config.get('username'),
            'target_password': target_config.get('password'),
            'target_table': target_config.get('table'),
            'target_sql': target_config.get('sql'),
            'target_file_type': target_config.get('file_type'),
            'target_delimiter': target_config.get('delimiter'),
            'target_file_path': target_config.get('file_path')
        }
        
        script_path = '/path/to/data_comparison.sh'
        command = build_command_with_args(script_path, args)
        
        success = execute_command(ssh_client, command)
        
        if success:
            report_path = f"/reports/comparison_{int(time.time())}.csv"
            socketio.emit('operation_complete', {
                'status': 'success',
                'operation': 'Data Comparison',
                'report_path': report_path
            })
        
    except Exception as e:
        emit_error(f"Data comparison failed: {str(e)}")

def execute_data_load(ssh_client, source_config, target_config, data):
    """Execute data load operation (legacy)"""
    try:
        emit_log("Starting data load operation...")
        
        args = {
            'source_type': source_config.get('type'),
            'source_host': source_config.get('host'),
            'source_database': source_config.get('database'),
            'source_username': source_config.get('username'),
            'source_password': source_config.get('password'),
            'source_table': source_config.get('table'),
            'target_type': target_config.get('type'),
            'target_host': target_config.get('host'),
            'target_database': target_config.get('database'),
            'target_username': target_config.get('username'),
            'target_password': target_config.get('password'),
            'target_table': target_config.get('table'),
            'batch_size': data.get('batchSize', '1000')
        }
        
        script_path = '/path/to/data_load.sh'
        command = build_command_with_args(script_path, args)
        
        success = execute_command(ssh_client, command)
        
        if success:
            socketio.emit('operation_complete', {
                'status': 'success',
                'operation': 'Data Load'
            })
        
    except Exception as e:
        emit_error(f"Data load failed: {str(e)}")

def execute_schema_generation(ssh_client, source_config, target_config, data):
    """Execute schema generation operation (legacy)"""
    try:
        emit_log("Starting schema generation operation...")
        
        args = {
            'source_database': source_config.get('database'),
            'source_host': source_config.get('host'),
            'source_username': source_config.get('username'),
            'source_password': source_config.get('password'),
            'target_database': target_config.get('database'),
            'schema_name': data.get('schemaName', ''),
            'output_format': data.get('outputFormat', 'DDL')
        }
        
        script_path = '/path/to/schema_generation.sh'
        command = build_command_with_args(script_path, args)
        
        success = execute_command(ssh_client, command)
        
        if success:
            socketio.emit('operation_complete', {
                'status': 'success',
                'operation': 'Schema Generation'
            })
        
    except Exception as e:
        emit_error(f"Schema generation failed: {str(e)}")

def execute_file_load(ssh_client, source_config, target_config, data):
    """Execute file load operation (legacy)"""
    try:
        emit_log("Starting file load operation...")
        
        args = {
            'file_type': source_config.get('file_type'),
            'file_path': source_config.get('file_path'),
            'target_database': target_config.get('database'),
            'target_table': target_config.get('table'),
            'target_host': target_config.get('host'),
            'target_username': target_config.get('username'),
            'target_password': target_config.get('password')
        }
        
        script_path = '/path/to/file_load.sh'
        command = build_command_with_args(script_path, args)
        
        success = execute_command(ssh_client, command)
        
        if success:
            socketio.emit('operation_complete', {
                'status': 'success',
                'operation': 'File Load'
            })
        
    except Exception as e:
        emit_error(f"File load failed: {str(e)}")

def execute_file_download(ssh_client, source_config, target_config, data):
    """Execute file download operation (legacy)"""
    try:
        emit_log("Starting file download operation...")
        
        args = {
            'source_path': source_config.get('file_path'),
            'target_path': target_config.get('file_path'),
            'file_format': data.get('exportFormat', 'CSV')
        }
        
        script_path = '/path/to/file_download.sh'
        command = build_command_with_args(script_path, args)
        
        success = execute_command(ssh_client, command)
        
        if success:
            socketio.emit('operation_complete', {
                'status': 'success',
                'operation': 'File Download'
            })
        
    except Exception as e:
        emit_error(f"File download failed: {str(e)}")

def extract_unix_config(data):
    """Extract Unix server configuration from request data (legacy)"""
    return {
        'host': data.get('comparisonUnixHost') or data.get('loadUnixHost') or data.get('schemaUnixHost') or data.get('fileloadUnixHost') or data.get('filedownloadUnixHost', 'localhost'),
        'port': data.get('comparisonUnixPort') or data.get('loadUnixPort') or data.get('schemaUnixPort') or data.get('fileloadUnixPort') or data.get('filedownloadUnixPort', '22'),
        'username': data.get('comparisonUnixUsername') or data.get('loadUnixUsername') or data.get('schemaUnixUsername') or data.get('fileloadUnixUsername') or data.get('filedownloadUnixUsername', 'user'),
        'password': data.get('comparisonUnixPassword') or data.get('loadUnixPassword') or data.get('schemaUnixPassword') or data.get('fileloadUnixPassword') or data.get('filedownloadUnixPassword', '')
    }

def extract_source_config(data, operation_type):
    """Extract source configuration from request data (legacy)"""
    prefix = get_prefix_for_operation(operation_type, 'source')
    
    return {
        'type': data.get(f'{prefix}Type'),
        'host': data.get(f'{prefix}Host'),
        'port': data.get(f'{prefix}Port'),
        'database': data.get(f'{prefix}Database'),
        'username': data.get(f'{prefix}Username'),
        'password': data.get(f'{prefix}Password'),
        'table': data.get(f'{prefix}TableName'),
        'sql': data.get(f'{prefix}SqlQuery'),
        'file_type': data.get(f'{prefix}FileType'),
        'delimiter': data.get(f'{prefix}Delimiter'),
        'file_path': data.get(f'{prefix}FilePath')
    }

def extract_target_config(data, operation_type):
    """Extract target configuration from request data (legacy)"""
    prefix = get_prefix_for_operation(operation_type, 'target')
    
    return {
        'type': data.get(f'{prefix}Type'),
        'host': data.get(f'{prefix}Host'),
        'port': data.get(f'{prefix}Port'),
        'database': data.get(f'{prefix}Database'),
        'username': data.get(f'{prefix}Username'),
        'password': data.get(f'{prefix}Password'),
        'table': data.get(f'{prefix}TableName'),
        'sql': data.get(f'{prefix}SqlQuery'),
        'file_type': data.get(f'{prefix}FileType'),
        'delimiter': data.get(f'{prefix}Delimiter'),
        'file_path': data.get(f'{prefix}FilePath')
    }

def get_prefix_for_operation(operation_type, side):
    """Get field prefix based on operation type and side (legacy)"""
    operation_prefixes = {
        'Data Comparison': 'comparison',
        'Data Load': 'load',
        'Schema Generation': 'schema',
        'File Load': 'fileload',
        'File Download': 'filedownload'
    }
    
    prefix = operation_prefixes.get(operation_type, 'comparison')
    return f"{prefix}{side.capitalize()}"

def emit_log(message):
    """Emit log message to connected clients (legacy WebSocket)"""
    timestamp = time.strftime('%H:%M:%S')
    log_message = f"[{timestamp}] {message}"
    socketio.emit('log', {'message': log_message})
    print(log_message)

def emit_error(error_message):
    """Emit error message to connected clients (legacy WebSocket)"""
    socketio.emit('operation_error', {'error': error_message})
    print(f"ERROR: {error_message}")

# ============================================================
# WEBSOCKET HANDLERS (legacy)
# ============================================================
@socketio.on('connect')
def handle_connect():
    print('Client connected')
    emit('connected', {'data': 'Connected to server'})

@socketio.on('disconnect')
def handle_disconnect():
    print('Client disconnected')


# ============================================================
# MISMATCH EXPLORER - In-memory storage for job results
# ============================================================
mismatch_jobs = {}
# Structure:
# {
#   job_id: {
#     'status': 'started' | 'completed' | 'failed',
#     'table': str,
#     'pkColumn': str,
#     'rules': dict,
#     'totalRecords': int,
#     'summaryRows': [],  # List of {pkValue, pkDisplay, mismatchCount, status}
#     'detailsCache': {},  # {pkValue: [field details]}
#     'created_at': datetime
#   }
# }


def parse_record_explorer_json(records_data, run_id=None):
     """
     Robustly parse records from the JSON report file format.
     Supports: JSON array, JSON Lines (NDJSON), single JSON object, and double-encoded JSON.
     Expected format: list of records with pk, pkDisplay, mismatchCount, status, fields
     
     Returns (summary_rows, details_cache, debug_info)
     """
     summary_rows = []
     details_cache = {}
     debug_info = {
         'detected_format': 'unknown',
         'raw_count': 0,
         'parsed_count': 0,
         'skipped_count': 0,
         'messages': []
     }
     
     # Step 1: Ensure records_data is a list of dicts
     if isinstance(records_data, str):
         # Double-encoded JSON safety
         try:
             records_data = json.loads(records_data)
             debug_info['messages'].append('Detected double-encoded JSON')
             if isinstance(records_data, str):
                 records_data = json.loads(records_data)
                 debug_info['messages'].append('Applied second decode')
         except json.JSONDecodeError as e:
             # Not valid JSON, try JSON Lines
             records_data = None
     
     # Step 2: Handle different JSON formats
     if records_data is None or (isinstance(records_data, str)):
         # Try JSON Lines format
         lines = records_data.strip().split('\n') if isinstance(records_data, str) else []
         if lines and len(lines) > 0:
             debug_info['detected_format'] = 'jsonlines'
             debug_info['messages'].append(f'Detected JSON Lines format with {len(lines)} lines')
             records_list = []
             for i, line in enumerate(lines):
                 line = line.strip()
                 if not line:
                     continue
                 try:
                     record = json.loads(line)
                     if isinstance(record, dict):
                         records_list.append(record)
                     else:
                         debug_info['skipped_count'] += 1
                         debug_info['messages'].append(f'Skipped line {i+1} (not a dict after parse)')
                 except json.JSONDecodeError:
                     debug_info['skipped_count'] += 1
                     debug_info['messages'].append(f'Skipped line {i+1} (invalid JSON)')
             records_data = records_list
     elif isinstance(records_data, dict):
         # Single JSON object, wrap in list
         debug_info['detected_format'] = 'single_object'
         debug_info['messages'].append('Detected single JSON object, wrapping in list')
         records_data = [records_data]
     elif isinstance(records_data, list):
         # Already a list, check format
         if len(records_data) > 0:
             if isinstance(records_data[0], dict):
                 debug_info['detected_format'] = 'array'
                 debug_info['messages'].append('Detected JSON array format')
             elif isinstance(records_data[0], str):
                 debug_info['detected_format'] = 'array_of_strings'
                 debug_info['messages'].append('Detected array of JSON strings, parsing each')
                 parsed_records = []
                 for i, item in enumerate(records_data):
                     try:
                         parsed_item = json.loads(item)
                         if isinstance(parsed_item, dict):
                             parsed_records.append(parsed_item)
                         else:
                             debug_info['skipped_count'] += 1
                             debug_info['messages'].append(f'Skipped index {i} (parsed to non-dict)')
                     except json.JSONDecodeError:
                         debug_info['skipped_count'] += 1
                         debug_info['messages'].append(f'Skipped index {i} (invalid JSON)')
                 records_data = parsed_records
     else:
         debug_info['messages'].append(f'Unexpected format: {type(records_data).__name__}')
         records_data = []
     
     # Step 3: Final validation and processing
     if not isinstance(records_data, list):
         records_data = []
     
     debug_info['raw_count'] = len(records_data)
     
     final_records = []
     for i, record in enumerate(records_data):
         if isinstance(record, dict):
             final_records.append(record)
         elif isinstance(record, str):
             # Try to parse string elements
             try:
                 parsed_record = json.loads(record)
                 if isinstance(parsed_record, dict):
                     final_records.append(parsed_record)
                 else:
                     debug_info['skipped_count'] += 1
                     debug_info['messages'].append(f'Skipped index {i} (re-parse not dict)')
             except json.JSONDecodeError:
                 debug_info['skipped_count'] += 1
                 debug_info['messages'].append(f'Skipped index {i} (cannot parse string)')
         else:
             debug_info['skipped_count'] += 1
             debug_info['messages'].append(f'Skipped index {i} (type: {type(record).__name__})')
     
     # Step 4: Build summary and details cache
     debug_info['parsed_count'] = len(final_records)
     for record in final_records:
         pk_display = record.get('pkDisplay', str(record.get('pk', 'UNKNOWN')))
         mismatch_count = record.get('mismatchCount', 0)
         status = record.get('status', 'MATCH')
         
         summary_rows.append({
             'pkValue': pk_display,
             'mismatchCount': mismatch_count,
             'status': status
         })
         
         # Store field details
         fields = record.get('fields', [])
         details_cache[pk_display] = fields
     
     # Log debug info if run_id provided
     if run_id:
         print(f"[Mismatch Explorer Fetch] Format: {debug_info['detected_format']}")
         print(f"[Mismatch Explorer Fetch] Raw records: {debug_info['raw_count']}")
         print(f"[Mismatch Explorer Fetch] Parsed records: {debug_info['parsed_count']}")
         print(f"[Mismatch Explorer Fetch] Skipped: {debug_info['skipped_count']}")
         for msg in debug_info['messages']:
             print(f"[Mismatch Explorer Fetch] {msg}")
     
     return summary_rows, details_cache, debug_info


def generate_mock_mismatch_data_from_records():
    """
    Generate mock mismatch data in the new JSON report format for testing.
    Returns list of records.
    """
    import random
    
    field_names = ['CUSTOMER_ID', 'FIRST_NAME', 'LAST_NAME', 'EMAIL', 'PHONE', 
                   'ADDRESS', 'CITY', 'STATE', 'ZIP_CODE', 'COUNTRY',
                   'CREATED_DATE', 'UPDATED_DATE', 'STATUS', 'BALANCE', 'ACCOUNT_TYPE']
    
    records = []
    num_records = random.randint(20, 100)
    
    for i in range(num_records):
        pk_obj = {
            'client_id': str(10 + i),
            'account_id': f'A{100 + i}',
            'batch_id': f'B{random.randint(1, 99)}'
        }
        pk_display = ' | '.join([f'{k}={v}' for k, v in pk_obj.items()])
        
        # Generate fields
        fields = []
        selected_fields = random.sample(field_names, random.randint(5, len(field_names)))
        mismatch_count = random.randint(0, 5)
        mismatch_indices = random.sample(range(len(selected_fields)), min(mismatch_count, len(selected_fields)))
        
        actual_mismatch_count = 0
        for j, field in enumerate(selected_fields):
            if j in mismatch_indices:
                source_val = f'SrcVal_{random.randint(100, 999)}'
                target_val = f'TgtVal_{random.randint(100, 999)}'
                field_status = 'MISMATCH'
                actual_mismatch_count += 1
            else:
                val = f'Val_{random.randint(100, 999)}'
                source_val = val
                target_val = val
                field_status = 'MATCH'
            
            fields.append({
                'fieldName': field,
                'sourceValue': source_val,
                'targetValue': target_val,
                'status': field_status
            })
        
        records.append({
            'pk': pk_obj,
            'pkDisplay': pk_display,
            'mismatchCount': actual_mismatch_count,
            'status': 'MISMATCH' if actual_mismatch_count > 0 else 'MATCH',
            'fields': fields
        })
    
    return records


# ============================================================
# MISMATCH EXPLORER ENDPOINTS
# ============================================================

@app.route('/api/mismatch-explorer/execute', methods=['POST'])
def mismatch_explorer_execute():
    """
    Execute Mismatch Explorer job via Paramiko on Unix.
    Returns run_id for log streaming.
    """
    try:
        data = request.get_json()
        
        # Extract Unix config
        unix_config = {
            'host': data.get('mismatchUnixHost', ''),
            'port': 22,
            'username': data.get('mismatchUnixUsername', ''),
            'password': data.get('mismatchUnixPassword', ''),
            'batch_id': data.get('mismatchBatchId', ''),
            'user_email': data.get('mismatchUserEmail', '')
        }
        
        # Validate required Unix fields
        if not unix_config['host']:
            return jsonify({'error': 'Unix Host is required'}), 400
        if not unix_config['username'] or not unix_config['password']:
            return jsonify({'error': 'Unix credentials are required'}), 400
        if not unix_config['batch_id']:
            return jsonify({'error': 'Batch ID is required'}), 400
        
        # Generate run ID
        run_id = f'mismatch_{uuid.uuid4().hex[:12]}'
        
        # Initialize run in registry
        run_registry[run_id] = {
            'status': 'PENDING',
            'logs': [],
            'tab': 'mismatch_explorer',
            'params': data,
            'unix_config': unix_config,
            'thread': None,
            'ssh_client': None,
            'stop_requested': False,
            'started_at': datetime.now()
        }
        
        # Start execution in background thread
        thread = threading.Thread(
            target=execute_mismatch_explorer_job,
            args=(run_id, data, unix_config)
        )
        run_registry[run_id]['thread'] = thread
        thread.start()
        
        return jsonify({
            'run_id': run_id,
            'status': 'started'
        })
        
    except Exception as e:
        print(f"Mismatch Explorer execute error: {str(e)}")
        return jsonify({'error': str(e)}), 500


def execute_mismatch_explorer_job(run_id, params, unix_config):
    """
    Background thread to execute Mismatch Explorer job via Paramiko.
    """
    try:
        run_registry[run_id]['status'] = 'RUNNING'
        add_log(run_id, "🚀 Starting Mismatch Explorer execution...")
        
        # Debug log: Mismatch Explorer Unix config received (mask password)
        add_log(run_id, f"Mismatch Explorer Unix config received: host={unix_config['host']}, user={unix_config['username']}, batch_id={unix_config['batch_id']}")
        
        # Log explicit connection details (NEVER log password)
        add_log(run_id, f"📡 Using Unix Host: {unix_config['host']}")
        add_log(run_id, f"👤 Using Unix Username: {unix_config['username']}")
        add_log(run_id, f"🔌 Using SSH Port: {unix_config['port']}")
        add_log(run_id, "🔐 Auth Method: password")
        add_log(run_id, "📡 Attempting SSH connection...")
        
        # Connect via SSH
        ssh_client = paramiko.SSHClient()
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        
        try:
            ssh_client.connect(
                hostname=unix_config['host'],
                port=unix_config['port'],
                username=unix_config['username'],
                password=unix_config['password'],
                timeout=30,
                allow_agent=False,
                look_for_keys=False
            )
            run_registry[run_id]['ssh_client'] = ssh_client
            add_log(run_id, f"✅ SSH connected successfully to {unix_config['host']}")
            
        except paramiko.AuthenticationException as auth_err:
            add_log(run_id, f"❌ SSH connection failed: Authentication failed - {str(auth_err)}")
            run_registry[run_id]['status'] = 'FAILED'
            run_registry[run_id]['finished_at'] = datetime.now()
            return
        except paramiko.SSHException as ssh_err:
            add_log(run_id, f"❌ SSH connection failed: SSH error - {str(ssh_err)}")
            run_registry[run_id]['status'] = 'FAILED'
            run_registry[run_id]['finished_at'] = datetime.now()
            return
        except Exception as conn_error:
            add_log(run_id, f"❌ SSH connection failed: {type(conn_error).__name__} - {str(conn_error)}")
            run_registry[run_id]['status'] = 'FAILED'
            run_registry[run_id]['finished_at'] = datetime.now()
            return
        
        # Resolve execution context using the unix_config we already have
        # Wrap unix_config as params for resolve_execution_context
        context_params = {
            'unix_host': unix_config['host'],
            'unixHost': unix_config['host'],
            'batch_id': unix_config['batch_id'],
            'batchId': unix_config['batch_id']
        }
        context = resolve_execution_context(context_params)
        if context.get('error'):
            add_log(run_id, f"❌ Context error: {context['error']}")
            run_registry[run_id]['status'] = 'FAILED'
            run_registry[run_id]['finished_at'] = datetime.now()
            ssh_client.close()
            return
        
        working_dir = context['working_dir']
        batch_id = context['batch_id']
        env = context['env']
        
        add_log(run_id, f"📂 Working directory: {working_dir}")
        add_log(run_id, f"🌍 Environment: {env}, Batch ID: {batch_id}")
        
        # Build the execution command
        script_name = 'run_mismatch_explorer.sh'
        command = f"cd {shlex.quote(working_dir)} && ./{script_name}"
        
        # Add active_tab parameter
        command += f" --active_tab=mismatch_explorer"
        add_log(run_id, f"🏷️ Active tab passed to Unix: mismatch_explorer")
        
        # Add Unix parameters explicitly
        command += f" --unix_host={shlex.quote(unix_config['host'].strip())}"
        command += f" --batch_id={shlex.quote(unix_config['batch_id'].strip())}"
        command += f" --unix_user={shlex.quote(unix_config['username'].strip())}"
        
        # Log Unix params confirmation
        add_log(run_id, f"Mismatch Explorer Unix params: unix_host={unix_config['host'].strip()}, batch_id={unix_config['batch_id'].strip()}, unix_user={unix_config['username'].strip()}")
        
        # Add parameters
        command += f" --run_id={shlex.quote(run_id)}"
        
        # Source config
        if params.get('mismatchSourceType'):
            command += f" --source_db_type={shlex.quote(params['mismatchSourceType'])}"
        if params.get('mismatchSourceHost'):
            command += f" --source_host={shlex.quote(params['mismatchSourceHost'])}"
        if params.get('mismatchSourcePort'):
            command += f" --source_port={shlex.quote(params['mismatchSourcePort'])}"
        if params.get('mismatchSourceDatabase'):
            command += f" --source_database={shlex.quote(params['mismatchSourceDatabase'])}"
        if params.get('mismatchSourceUsername'):
            command += f" --source_username={shlex.quote(params['mismatchSourceUsername'])}"
        if params.get('mismatchSourcePassword'):
            command += f" --source_password={shlex.quote(params['mismatchSourcePassword'])}"
        if params.get('mismatchSourceTableName'):
            command += f" --source_table={shlex.quote(params['mismatchSourceTableName'])}"
        if params.get('mismatchSourcePrimaryKey'):
            command += f" --source_pk={shlex.quote(params['mismatchSourcePrimaryKey'])}"
        if params.get('mismatchSourceSqlQuery'):
            command += f" --source_sql={shlex.quote(params['mismatchSourceSqlQuery'])}"
        
        # Target config
        if params.get('mismatchTargetType'):
            command += f" --target_db_type={shlex.quote(params['mismatchTargetType'])}"
        if params.get('mismatchTargetHost'):
            command += f" --target_host={shlex.quote(params['mismatchTargetHost'])}"
        if params.get('mismatchTargetPort'):
            command += f" --target_port={shlex.quote(params['mismatchTargetPort'])}"
        if params.get('mismatchTargetDatabase'):
            command += f" --target_database={shlex.quote(params['mismatchTargetDatabase'])}"
        if params.get('mismatchTargetUsername'):
            command += f" --target_username={shlex.quote(params['mismatchTargetUsername'])}"
        if params.get('mismatchTargetPassword'):
            command += f" --target_password={shlex.quote(params['mismatchTargetPassword'])}"
        if params.get('mismatchTargetTableName'):
            command += f" --target_table={shlex.quote(params['mismatchTargetTableName'])}"
        if params.get('mismatchTargetPrimaryKey'):
            command += f" --target_pk={shlex.quote(params['mismatchTargetPrimaryKey'])}"
        if params.get('mismatchTargetSqlQuery'):
            command += f" --target_sql={shlex.quote(params['mismatchTargetSqlQuery'])}"
        
        # Rules
        rules = params.get('rules', {})
        if rules:
            command += f" --trim_spaces={'true' if rules.get('trim_spaces') else 'false'}"
            command += f" --ignore_case={'true' if rules.get('ignore_case') else 'false'}"
            command += f" --null_equals_empty={'true' if rules.get('null_equals_empty') else 'false'}"
            command += f" --numeric_tolerance={rules.get('numeric_tolerance', 0)}"
            command += f" --time_zone={shlex.quote(rules.get('time_zone', 'N/A'))}"
        
        add_log(run_id, f"▶️ Executing script: {script_name}")
        add_log(run_id, f"📋 Command: {command}")
        
        # Execute command
        stdin, stdout, stderr = ssh_client.exec_command(command, get_pty=True)
        
        # Stream output
        for line in iter(stdout.readline, ''):
            if run_registry[run_id].get('stop_requested'):
                add_log(run_id, "⚠️ Execution stopped by user")
                break
            line = line.strip()
            if line:
                add_log(run_id, line)
        
        # Check for errors
        error_output = stderr.read().decode('utf-8', errors='replace').strip()
        if error_output:
            add_log(run_id, f"⚠️ Stderr: {error_output}")
        
        exit_code = stdout.channel.recv_exit_status()
        
        if exit_code == 0:
            add_log(run_id, f"✅ Script execution completed (rc=0). Fetching /tmp/{run_id}_record_explorer.json...")
            
            # Immediately fetch report in the SAME SSH session (OTP safe)
            report_fetched = False
            json_content = None
            json_filename = f'{run_id}_record_explorer.json'
            remote_path = f'/tmp/{json_filename}'
            
            try:
                sftp = ssh_client.open_sftp()
                
                # Retry stat() up to 5 times with 1s sleep (file write may finish slightly later)
                file_found = False
                for attempt in range(1, 6):
                    try:
                        sftp.stat(remote_path)
                        file_found = True
                        break
                    except FileNotFoundError:
                        add_log(run_id, f"Waiting for report file... attempt {attempt}/5")
                        time.sleep(1)
                
                # If not found by stat, try ls to locate alternative filename
                if not file_found:
                    add_log(run_id, f"❌ Report file not found after 5 retries: {remote_path}")
                    try:
                        ls_stdin, ls_stdout, ls_stderr = ssh_client.exec_command(
                            f'ls -1 /tmp | grep -E "{run_id}.*record_explorer\\.json"'
                        )
                        ls_output = ls_stdout.read().decode('utf-8').strip()
                        if ls_output:
                            # Use first match
                            alt_name = ls_output.split('\n')[0].strip()
                            remote_path = f'/tmp/{alt_name}'
                            add_log(run_id, f"✅ Found alternative file via ls: {remote_path}")
                            file_found = True
                        else:
                            # Log what's in /tmp for debugging
                            ls2_stdin, ls2_stdout, ls2_stderr = ssh_client.exec_command(
                                f'ls -1 /tmp | grep -i "{run_id}" | head -20'
                            )
                            ls2_output = ls2_stdout.read().decode('utf-8').strip()
                            add_log(run_id, f"Report not found. Listing /tmp matches: {ls2_output if ls2_output else '(none)'}")
                    except Exception as ls_err:
                        add_log(run_id, f"⚠️ ls check failed: {str(ls_err)}")
                
                if file_found:
                    try:
                        with sftp.file(remote_path, 'r') as f:
                            json_content = f.read().decode('utf-8')
                        report_fetched = True
                    except Exception as read_err:
                        add_log(run_id, f"❌ Error reading report file: {str(read_err)}")
                
                sftp.close()
            except Exception as sftp_err:
                add_log(run_id, f"❌ SFTP error: {str(sftp_err)}")
            
            if report_fetched and json_content:
                # Robust JSON parsing: support JSON array/object OR JSON Lines
                records = None
                try:
                    records = json.loads(json_content)
                except json.JSONDecodeError as je:
                    if 'Extra data' in str(je) or 'extra data' in str(je).lower():
                        # Treat as JSON Lines
                        try:
                            lines = [l.strip() for l in json_content.strip().split('\n') if l.strip()]
                            records = [json.loads(line) for line in lines]
                            add_log(run_id, f"Detected JSON Lines format. Parsed {len(records)} records.")
                        except Exception as jl_err:
                            add_log(run_id, f"❌ JSON Lines parsing also failed: {str(jl_err)}")
                            add_log(run_id, f"📝 Raw content (first 500 chars): {json_content[:500]}")
                    else:
                        add_log(run_id, f"❌ JSON parse error: {str(je)}")
                        add_log(run_id, f"📝 Raw content (first 500 chars): {json_content[:500]}")
                
                if records is not None:
                    # Ensure records is a list
                    if isinstance(records, dict):
                        records = [records]
                    
                    summary_rows, details_cache, debug_info = parse_record_explorer_json(records, run_id)
                    add_log(run_id, f"📄 Report fetched successfully. Parsed records: {len(summary_rows)}")
                    
                    job_id = f'job_{run_id}'
                    mismatch_jobs[job_id] = {
                        'status': 'completed',
                        'totalRecords': len(summary_rows),
                        'summaryRows': summary_rows,
                        'detailsCache': details_cache,
                        'created_at': datetime.now()
                    }
                    run_registry[run_id]['cached_job_id'] = job_id
                    run_registry[run_id]['has_results'] = True
                    run_registry[run_id]['status'] = 'SUCCESS'
                    add_log(run_id, "💾 Results cached for Fetch & Compare (no SSH needed).")
                else:
                    add_log(run_id, "❌ Failed to parse report JSON. No results cached.")
                    run_registry[run_id]['has_results'] = False
                    run_registry[run_id]['status'] = 'FAILED'
            elif not report_fetched:
                add_log(run_id, "❌ Report file could not be fetched from /tmp.")
                run_registry[run_id]['has_results'] = False
                run_registry[run_id]['status'] = 'FAILED'
        else:
            run_registry[run_id]['status'] = 'FAILED'
            add_log(run_id, f"❌ Execution failed with exit code: {exit_code}")
        
        # Close SSH only AFTER report fetch attempts are complete
        ssh_client.close()
        add_log(run_id, "🔌 SSH connection closed")
        
    except Exception as e:
        add_log(run_id, f"❌ Error: {str(e)}")
        run_registry[run_id]['status'] = 'FAILED'
        print(f"Error in execute_mismatch_explorer_job: {str(e)}")
    
    run_registry[run_id]['finished_at'] = datetime.now()


@app.route('/api/mismatch-explorer/status', methods=['GET'])
def mismatch_explorer_status():
    """Lightweight status endpoint for Mismatch Explorer polling."""
    run_id = request.args.get('runId', '')
    if not run_id or run_id not in run_registry:
        return jsonify({'error': 'Run ID not found'}), 404
    entry = run_registry[run_id]
    status = entry.get('status', 'PENDING')
    has_results = entry.get('has_results', False)
    is_running = status in ('PENDING', 'RUNNING')
    return jsonify({
        'runId': run_id,
        'status': status,
        'hasResults': has_results,
        'isRunning': is_running
    })


@app.route('/api/mismatch-explorer/fetch/<run_id>', methods=['GET'])
def mismatch_explorer_fetch(run_id):
    """
    Return cached results from Execute (no SSH reconnect needed - OTP safe).
    """
    try:
        if run_id not in run_registry:
            return jsonify({'error': 'Invalid or expired run ID'}), 404
        
        run_info = run_registry[run_id]
        
        # Check for cached results from Execute
        cached_job_id = run_info.get('cached_job_id')
        if cached_job_id and cached_job_id in mismatch_jobs:
            add_log(run_id, "📥 Returning cached results (no SSH needed).")
            job = mismatch_jobs[cached_job_id]
            return jsonify({
                'jobId': cached_job_id,
                'status': 'completed',
                'totalRecords': job['totalRecords']
            })
        
        # No cached results available
        add_log(run_id, "❌ Results not available. Please re-run Execute.")
        return jsonify({'error': 'Results not available. Please re-run Execute.'}), 400
        
    except Exception as e:
        print(f"Mismatch Explorer fetch error: {str(e)}")
        add_log(run_id, f"❌ Fetch error: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/mismatch-explorer/summary', methods=['GET'])
def mismatch_explorer_summary():
    """
    Get paginated summary of comparison results.
    """
    try:
        job_id = request.args.get('jobId', '')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('pageSize', 50))
        
        if not job_id or job_id not in mismatch_jobs:
            return jsonify({'error': 'Invalid or expired job ID'}), 404
        
        job = mismatch_jobs[job_id]
        summary_rows = job['summaryRows']
        total_records = len(summary_rows)
        
        # Paginate
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_rows = summary_rows[start_idx:end_idx]
        
        return jsonify({
            'totalRecords': total_records,
            'page': page,
            'pageSize': page_size,
            'totalPages': (total_records + page_size - 1) // page_size,
            'rows': paginated_rows
        })
        
    except Exception as e:
        print(f"Mismatch Explorer summary error: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/mismatch-explorer/details', methods=['GET'])
def mismatch_explorer_details():
    """
    Get field-level details for a specific PK value.
    """
    try:
        job_id = request.args.get('jobId', '')
        pk_value = request.args.get('pkValue', '')
        
        if not job_id or job_id not in mismatch_jobs:
            return jsonify({'error': 'Invalid or expired job ID'}), 404
        
        if not pk_value:
            return jsonify({'error': 'PK value is required'}), 400
        
        job = mismatch_jobs[job_id]
        details_cache = job['detailsCache']
        
        if pk_value not in details_cache:
            return jsonify({'error': f'No details found for PK: {pk_value}'}), 404
        
        return jsonify({
            'pkValue': pk_value,
            'fields': details_cache[pk_value]
        })
        
    except Exception as e:
        print(f"Mismatch Explorer details error: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/mismatch-explorer/export', methods=['GET'])
def mismatch_explorer_export():
    """
    Export single record comparison as CSV.
    """
    try:
        job_id = request.args.get('jobId', '')
        pk_value = request.args.get('pkValue', '')
        
        if not job_id or job_id not in mismatch_jobs:
            return jsonify({'error': 'Invalid or expired job ID'}), 404
        
        if not pk_value:
            return jsonify({'error': 'PK value is required'}), 400
        
        job = mismatch_jobs[job_id]
        details_cache = job['detailsCache']
        
        if pk_value not in details_cache:
            return jsonify({'error': f'No details found for PK: {pk_value}'}), 404
        
        fields = details_cache[pk_value]
        
        # Build CSV content
        csv_lines = ['Field Name,Source Value,Target Value,Status']
        for field in fields:
            # Escape values for CSV
            field_name = field['fieldName'].replace('"', '""')
            source_val = str(field['sourceValue']).replace('"', '""')
            target_val = str(field['targetValue']).replace('"', '""')
            status = field['status']
            csv_lines.append(f'"{field_name}","{source_val}","{target_val}","{status}"')
        
        csv_content = '\n'.join(csv_lines)
        
        # Create response with CSV download
        safe_pk = pk_value.replace('|', '_').replace('/', '_').replace('\\', '_').replace(' ', '_')[:50]
        filename = f'mismatch_{safe_pk}.csv'
        
        return Response(
            csv_content,
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
        
    except Exception as e:
        print(f"Mismatch Explorer export error: {str(e)}")
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    socketio.run(app, debug=True, host='0.0.0.0', port=5000)
